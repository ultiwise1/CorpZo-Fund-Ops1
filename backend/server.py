"""CorpZo Debt CRM — main FastAPI server."""
from __future__ import annotations
from fastapi import FastAPI, APIRouter, Request, Response, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import Response as FastResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, List, Any, Dict
import os, uuid, logging, io, asyncio, re

from auth import fetch_session_data, set_session_cookie, clear_session_cookie, get_current_user
from storage import init_storage, put_object, get_object, APP_NAME
from uids import gen_uid
from seed import seed_demo, seed_supplemental
from notify import notify, send_email

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
log = logging.getLogger("corpzo")

app = FastAPI(title="CorpZo Debt CRM")
api = APIRouter(prefix="/api")


# ----------- audit -----------
def _sanitize(v):
    """Strip Mongo ObjectId values so audit payloads stay JSON-serialisable."""
    if isinstance(v, dict):
        return {k: _sanitize(x) for k, x in v.items() if k != "_id"}
    if isinstance(v, list):
        return [_sanitize(x) for x in v]
    if isinstance(v, ObjectId):
        return str(v)
    return v


async def audit(actor: dict, entity_type: str, entity_id: str, action: str,
                before: Optional[dict] = None, after: Optional[dict] = None):
    before = _sanitize(before)
    after = _sanitize(after)
    await db.audit_logs.insert_one({
        "audit_id": f"aud_{uuid.uuid4().hex[:12]}",
        "actor_uid": actor.get("user_id") if actor else "system",
        "actor_name": actor.get("name") if actor else "System",
        "entity_type": entity_type, "entity_id": entity_id,
        "action": action, "before": before, "after": after,
        "at": datetime.now(timezone.utc).isoformat(),
    })


async def push_activity(entity_type: str, entity_id: str, kind: str, author: dict, summary: str, details: dict = None, **kw):
    await db.activities.insert_one({
        "activity_id": f"act_{uuid.uuid4().hex[:10]}",
        "entity_type": entity_type, "entity_id": entity_id,
        "kind": kind, "author_uid": author.get("user_id") if author else "system",
        "author_name": author.get("name") if author else "System",
        "summary": summary, "details": details or {},
        "duration_sec": kw.get("duration_sec"),
        "outcome": kw.get("outcome"), "next_followup": kw.get("next_followup"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    })


# ============== AUTH ==============
@api.post("/auth/session")
async def auth_session(request: Request, response: Response):
    body = await request.json()
    session_id = body.get("session_id")
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")

    data = await fetch_session_data(session_id)
    email = data.get("email")
    name = data.get("name", "")
    picture = data.get("picture", "")
    session_token = data.get("session_token")

    if not email or not session_token:
        raise HTTPException(status_code=401, detail="Invalid Google session data")

    existing = await db.users.find_one({"email": email}, {"_id": 0})
    now = datetime.now(timezone.utc)
    if existing:
        user_id = existing["user_id"]
        await db.users.update_one({"user_id": user_id}, {"$set": {
            "name": name or existing.get("name"),
            "picture": picture or existing.get("picture"),
            "last_login": now.isoformat(),
        }})
        role = existing.get("role", "sales_agent")
    else:
        # New user — first-ever signup gets super_admin, else default sales_agent
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        role = "super_admin" if email == os.environ.get("OWNER_EMAIL") else "customer"
        emp_uid = await gen_uid(db, "employee")
        await db.users.insert_one({
            "user_id": user_id, "email": email, "name": name, "picture": picture,
            "role": role, "employee_uid": emp_uid, "active": True,
            "created_at": now.isoformat(), "last_login": now.isoformat(),
        })
        await db.employees.insert_one({
            "employee_uid": emp_uid, "user_id": user_id, "email": email, "name": name,
            "role": role, "manager_uid": None,
            "joining_date": now.isoformat(),
            "ctc_monthly": 0, "target_multiplier": 3.0,
            "revenue_target": 0, "disbursement_target": 0,
            "login_target": 0, "sanction_target": 0,
            "active": True, "created_at": now.isoformat(),
        })

    await db.user_sessions.insert_one({
        "user_id": user_id, "session_token": session_token,
        "expires_at": (now + timedelta(days=7)).isoformat(),
        "ip": request.client.host if request.client else None,
        "created_at": now.isoformat(),
    })
    set_session_cookie(response, session_token)
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    await audit(user_doc, "auth", user_id, "login")
    return user_doc


@api.get("/auth/me")
async def auth_me(request: Request):
    user = await get_current_user(request, db)
    return user


@api.post("/auth/logout")
async def auth_logout(request: Request, response: Response):
    token = request.cookies.get("session_token")
    if token:
        await db.user_sessions.delete_one({"session_token": token})
    clear_session_cookie(response)
    return {"ok": True}


# ============== helpers ==============
async def require_user(request: Request):
    return await get_current_user(request, db)


# ---------- RBAC scoping ----------
FULL_ACCESS_ROLES = {"super_admin", "business_head", "finance", "compliance", "operations"}
CREDIT_ROLES = {"credit_head", "credit_analyst"}
MANAGER_ROLES = {"sales_manager", "channel_manager"}
AGENT_ROLES = {"sales_agent"}
PARTNER_ROLES = {"channel_partner"}


async def team_employee_uids(manager_uid: str) -> list:
    subs = await db.employees.find({"manager_uid": manager_uid}, {"_id": 0, "employee_uid": 1}).to_list(100)
    return [manager_uid] + [s["employee_uid"] for s in subs]


async def scope_query(user: dict, entity: str) -> dict:
    """Return an additional Mongo query filter based on the user's role.
    Empty {} means unrestricted access."""
    role = user.get("role", "sales_agent")
    if role in FULL_ACCESS_ROLES:
        return {}
    emp = user.get("employee_uid")
    partner_uid = user.get("partner_uid")

    if role in PARTNER_ROLES:
        if not partner_uid:
            return {"channel_partner_uid": "__none__"}  # deny-by-default
        if entity in ("leads", "cases", "clients"):
            return {"channel_partner_uid": partner_uid}
        return {"channel_partner_uid": partner_uid}

    if role in AGENT_ROLES:
        if entity == "leads": return {"assigned_to": emp}
        if entity == "cases": return {"sales_owner": emp}
        if entity == "clients": return {"relationship_manager": emp}
        if entity == "tasks": return {"owner_uid": emp}
        return {}

    if role in MANAGER_ROLES:
        team = await team_employee_uids(emp)
        if entity == "leads": return {"assigned_to": {"$in": team}}
        if entity == "cases": return {"sales_owner": {"$in": team}}
        if entity == "clients": return {"relationship_manager": {"$in": team}}
        if entity == "tasks": return {"owner_uid": {"$in": team}}
        return {}

    if role in CREDIT_ROLES:
        if entity == "cases": return {"credit_owner": emp} if role == "credit_analyst" else {}
        return {}
    return {}


def sanitize_partner_case(case: dict) -> dict:
    """Strip confidential credit notes for partner viewers."""
    hide = {"credit_owner", "expected_revenue", "actual_revenue"}
    return {k: v for k, v in case.items() if k not in hide}


# ============== PERMISSION SYSTEM (per-user overrides on top of role RBAC) ==============
PERMISSION_KEYS = {
    "release_commissions":   "Release monthly CP commission payout batches",
    "mark_payout_paid":      "Mark payout batches as PAID (closes CP + incentive items)",
    "create_partner":        "Create a new channel partner",
    "edit_partner_target":   "Edit a channel partner's monthly disbursement target",
    "publish_cam_template":  "Publish and delete shared CAM templates",
    "manage_users":          "Manage users, roles, and per-user permissions",
}

DEFAULT_ROLE_PERMISSIONS = {
    "super_admin":     set(PERMISSION_KEYS.keys()),
    "business_head":   {"release_commissions", "mark_payout_paid", "create_partner", "edit_partner_target", "publish_cam_template"},
    "finance":         {"release_commissions", "mark_payout_paid"},
    "channel_manager": {"create_partner", "edit_partner_target"},
    "credit_head":     {"publish_cam_template"},
    "credit_analyst":  {"publish_cam_template"},
}


def effective_permissions(user: dict) -> set:
    role = user.get("role", "")
    base = set(DEFAULT_ROLE_PERMISSIONS.get(role, set()))
    grants = set(user.get("permissions_grants") or [])
    revokes = set(user.get("permissions_revokes") or [])
    return (base | grants) - revokes


def has_permission(user: dict, key: str) -> bool:
    return key in effective_permissions(user)


def require_permission(user: dict, key: str):
    if not has_permission(user, key):
        raise HTTPException(403, f"Missing permission: {key}")


@api.get("/permissions/keys")
async def list_permission_keys(request: Request):
    await require_user(request)
    return [{"key": k, "label": v} for k, v in PERMISSION_KEYS.items()]


@api.get("/me/permissions")
async def my_permissions(request: Request):
    user = await require_user(request)
    return {
        "role": user.get("role"),
        "permissions": sorted(effective_permissions(user)),
        "grants": user.get("permissions_grants") or [],
        "revokes": user.get("permissions_revokes") or [],
    }


@api.get("/admin/users/{user_id}/permissions")
async def get_user_permissions(user_id: str, request: Request):
    actor = await require_user(request)
    require_permission(actor, "manage_users")
    u = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not u:
        raise HTTPException(404, "User not found")
    return {
        "user_id": u["user_id"], "role": u.get("role"),
        "permissions": sorted(effective_permissions(u)),
        "grants": u.get("permissions_grants") or [],
        "revokes": u.get("permissions_revokes") or [],
    }


@api.put("/admin/users/{user_id}/permissions")
async def set_user_permissions(user_id: str, request: Request):
    actor = await require_user(request)
    require_permission(actor, "manage_users")
    body = await request.json()
    grants = [k for k in (body.get("grants") or []) if k in PERMISSION_KEYS]
    revokes = [k for k in (body.get("revokes") or []) if k in PERMISSION_KEYS]
    before = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not before:
        raise HTTPException(404, "User not found")
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"permissions_grants": grants, "permissions_revokes": revokes}},
    )
    after = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    await audit(actor, "user", user_id, "permissions_updated", before, after)
    return {
        "user_id": user_id, "role": after.get("role"),
        "permissions": sorted(effective_permissions(after)),
        "grants": grants, "revokes": revokes,
    }


async def list_collection(coll_name: str, query: dict = None, sort_field: str = "created_at", limit: int = 500):
    query = query or {}
    cur = db[coll_name].find(query, {"_id": 0}).sort(sort_field, -1).limit(limit)
    return await cur.to_list(limit)


# ============== DASHBOARD ==============
@api.get("/dashboard/summary")
async def dashboard_summary(request: Request):
    user = await require_user(request)
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()

    leads_total = await db.leads.count_documents({})
    leads_today = await db.leads.count_documents({"created_at": {"$gte": today_start}})
    clients_total = await db.clients.count_documents({})
    cases_total = await db.cases.count_documents({})
    sanctioned = await db.cases.count_documents({"stage": {"$in": ["sanctioned", "disbursement_pending", "partially_disbursed", "fully_disbursed"]}})
    disbursed_cases = await db.cases.count_documents({"stage": {"$in": ["partially_disbursed", "fully_disbursed"]}})

    pipeline = list(await db.cases.aggregate([
        {"$group": {
            "_id": None,
            "requested": {"$sum": "$requirement"},
            "sanctioned": {"$sum": "$sanctioned_amount"},
            "disbursed": {"$sum": "$disbursed_amount"},
            "revenue": {"$sum": "$actual_revenue"},
            "expected_revenue": {"$sum": "$expected_revenue"},
        }},
    ]).to_list(1))
    agg = pipeline[0] if pipeline else {}

    # stage funnel
    stage_agg = await db.cases.aggregate([
        {"$group": {"_id": "$stage", "count": {"$sum": 1}, "value": {"$sum": "$requirement"}}}
    ]).to_list(50)
    funnel = [{"stage": s["_id"], "count": s["count"], "value": s["value"]} for s in stage_agg]

    # invoice collections
    inv_agg = await db.invoices.aggregate([
        {"$group": {"_id": "$status", "amount": {"$sum": "$amount"}, "count": {"$sum": 1}}}
    ]).to_list(20)

    # top lenders
    lender_agg = await db.applications.aggregate([
        {"$group": {"_id": "$lender_id", "apps": {"$sum": 1},
                    "sanctioned": {"$sum": {"$cond": [{"$eq": ["$status", "sanctioned"]}, 1, 0]}}}},
        {"$sort": {"apps": -1}}, {"$limit": 6},
    ]).to_list(6)

    # overdue tasks
    now_iso = datetime.now(timezone.utc).isoformat()
    overdue_tasks = await db.tasks.count_documents({"status": {"$ne": "done"}, "due_date": {"$lt": now_iso}})

    return {
        "user": {"user_id": user["user_id"], "name": user["name"], "role": user["role"]},
        "kpis": {
            "leads_total": leads_total, "leads_today": leads_today,
            "clients": clients_total, "cases": cases_total,
            "sanctioned": sanctioned, "disbursed_cases": disbursed_cases,
            "requested_amount": agg.get("requested", 0),
            "sanctioned_amount": agg.get("sanctioned", 0),
            "disbursed_amount": agg.get("disbursed", 0),
            "revenue_booked": agg.get("expected_revenue", 0),
            "revenue_collected": agg.get("revenue", 0),
            "overdue_tasks": overdue_tasks,
        },
        "funnel": funnel,
        "invoices_summary": [{"status": s["_id"], "amount": s["amount"], "count": s["count"]} for s in inv_agg],
        "top_lenders": [{"lender_id": l["_id"], "apps": l["apps"], "sanctioned": l["sanctioned"]} for l in lender_agg],
    }


# ============== LEADS ==============
@api.get("/leads")
async def get_leads(request: Request, stage: Optional[str] = None, priority: Optional[str] = None,
                    assigned: Optional[str] = None, source: Optional[str] = None, q: Optional[str] = None):
    user = await require_user(request)
    query = await scope_query(user, "leads")
    if stage: query["stage"] = stage
    if priority: query["priority"] = priority
    if assigned: query["assigned_to"] = assigned
    if source: query["source"] = source
    if q:
        query["$or"] = [
            {"name": {"$regex": q, "$options": "i"}},
            {"company": {"$regex": q, "$options": "i"}},
            {"mobile": {"$regex": q}},
            {"email": {"$regex": q, "$options": "i"}},
            {"lead_uid": {"$regex": q, "$options": "i"}},
        ]
    return await list_collection("leads", query)


@api.post("/leads")
async def create_lead(request: Request):
    user = await require_user(request)
    body = await request.json()
    lead_uid = await gen_uid(db, "lead")
    now_iso = datetime.now(timezone.utc).isoformat()
    doc = {
        "lead_uid": lead_uid, "source": body.get("source", "manual"),
        "source_detail": body.get("source_detail", ""), "campaign": body.get("campaign", ""),
        "referral": body.get("referral", ""), "channel_partner_uid": body.get("channel_partner_uid"),
        "assigned_to": body.get("assigned_to") or user.get("employee_uid"),
        "original_owner": body.get("assigned_to") or user.get("employee_uid"),
        "borrower_type": body.get("borrower_type", "business"),
        "name": body.get("name", ""), "company": body.get("company", ""),
        "mobile": body.get("mobile", ""), "email": body.get("email", ""),
        "city": body.get("city", ""), "state": body.get("state", ""),
        "product": body.get("product", "business_loan"),
        "approx_requirement": float(body.get("approx_requirement", 0)),
        "notes": body.get("notes", ""), "stage": "new_lead",
        "priority": body.get("priority", "warm"), "probability": int(body.get("probability", 30)),
        "expected_closure": body.get("expected_closure"), "rejection_reason": None,
        "client_uid": None, "converted": False, "duplicate_of": None,
        "created_at": now_iso,
    }
    await db.leads.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "lead", lead_uid, "created", None, doc)
    await push_activity("lead", lead_uid, "note", user, f"Lead created: {doc['name']}")
    return doc


@api.get("/leads/{lead_uid}")
async def get_lead(lead_uid: str, request: Request):
    await require_user(request)
    lead = await db.leads.find_one({"lead_uid": lead_uid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    activities = await db.activities.find({"entity_type": "lead", "entity_id": lead_uid}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return {"lead": lead, "activities": activities}


@api.patch("/leads/{lead_uid}")
async def update_lead(lead_uid: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    before = await db.leads.find_one({"lead_uid": lead_uid}, {"_id": 0})
    if not before:
        raise HTTPException(404, "Not found")
    allowed = {"stage", "priority", "probability", "assigned_to", "rejection_reason", "notes",
               "approx_requirement", "product", "expected_closure", "mobile", "email", "city", "state", "company", "name"}
    updates = {k: v for k, v in body.items() if k in allowed}
    if updates:
        await db.leads.update_one({"lead_uid": lead_uid}, {"$set": updates})
        after = await db.leads.find_one({"lead_uid": lead_uid}, {"_id": 0})
        await audit(user, "lead", lead_uid, "updated", before, after)
        if "stage" in updates and updates["stage"] != before.get("stage"):
            await push_activity("lead", lead_uid, "status_change", user,
                                f"Stage: {before.get('stage')} → {updates['stage']}")
        return after
    return before


@api.post("/leads/{lead_uid}/convert")
async def convert_lead(lead_uid: str, request: Request):
    user = await require_user(request)
    lead = await db.leads.find_one({"lead_uid": lead_uid}, {"_id": 0})
    if not lead: raise HTTPException(404, "Not found")
    if lead.get("converted"): raise HTTPException(400, "Already converted")

    client_uid = await gen_uid(db, "client")
    now_iso = datetime.now(timezone.utc).isoformat()
    client_doc = {
        "client_uid": client_uid, "name": lead["name"], "borrower_type": lead["borrower_type"],
        "company": lead["company"], "pan": "", "cin": "", "gstin": "",
        "mobile": lead["mobile"], "email": lead["email"],
        "city": lead["city"], "state": lead["state"],
        "industry": "", "constitution": "", "incorporation_date": None,
        "relationship_manager": lead["assigned_to"],
        "channel_partner_uid": lead.get("channel_partner_uid"),
        "source": lead["source"], "tags": [], "lead_uid": lead_uid, "created_at": now_iso,
    }
    await db.clients.insert_one(client_doc)
    client_doc.pop("_id", None)

    case_uid = await gen_uid(db, "case")
    case_doc = {
        "case_uid": case_uid, "client_uid": client_uid,
        "product": lead["product"], "requirement": lead["approx_requirement"],
        "purpose": "", "tenure_months": 60, "security": "", "geography": lead["state"],
        "expected_roi": 11.0, "urgency": "normal",
        "sales_owner": lead["assigned_to"], "credit_owner": None,
        "channel_partner_uid": lead.get("channel_partner_uid"),
        "source": lead["source"], "stage": "qualified",
        "expected_closure": lead.get("expected_closure"),
        "expected_revenue": lead["approx_requirement"] * 0.012, "actual_revenue": 0,
        "sanctioned_amount": 0, "disbursed_amount": 0,
        "lead_uid": lead_uid, "documentation_pct": 0, "created_at": now_iso,
    }
    await db.cases.insert_one(case_doc)
    case_doc.pop("_id", None)
    await db.leads.update_one({"lead_uid": lead_uid}, {"$set": {"converted": True, "client_uid": client_uid, "stage": "qualified"}})
    await audit(user, "lead", lead_uid, "converted", None, {"client_uid": client_uid, "case_uid": case_uid})
    await push_activity("lead", lead_uid, "status_change", user, f"Converted to client {client_uid} / case {case_uid}")
    return {"client": client_doc, "case": case_doc}


@api.post("/leads/{lead_uid}/assign")
async def assign_lead(lead_uid: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    emp = body.get("employee_uid")
    lead = await db.leads.find_one({"lead_uid": lead_uid}, {"_id": 0})
    if not lead: raise HTTPException(404, "Not found")
    await db.leads.update_one({"lead_uid": lead_uid}, {"$set": {"assigned_to": emp, "stage": "assigned"}})
    await audit(user, "lead", lead_uid, "assigned", {"assigned_to": lead.get("assigned_to")}, {"assigned_to": emp})
    await push_activity("lead", lead_uid, "note", user, f"Assigned to {emp}")
    return await db.leads.find_one({"lead_uid": lead_uid}, {"_id": 0})


# Activities on any entity
@api.post("/activities")
async def add_activity(request: Request):
    user = await require_user(request)
    body = await request.json()
    doc = {
        "activity_id": f"act_{uuid.uuid4().hex[:10]}",
        "entity_type": body["entity_type"], "entity_id": body["entity_id"],
        "kind": body.get("kind", "note"),
        "author_uid": user["user_id"], "author_name": user["name"],
        "summary": body.get("summary", ""), "details": body.get("details", {}),
        "duration_sec": body.get("duration_sec"),
        "outcome": body.get("outcome"),
        "next_followup": body.get("next_followup"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.activities.insert_one(doc)
    doc.pop("_id", None)
    return doc


# ============== CLIENTS ==============
@api.get("/clients")
async def get_clients(request: Request, q: Optional[str] = None):
    user = await require_user(request)
    query = await scope_query(user, "clients")
    if q:
        query["$or"] = [
            {"name": {"$regex": q, "$options": "i"}},
            {"company": {"$regex": q, "$options": "i"}},
            {"mobile": {"$regex": q}}, {"email": {"$regex": q, "$options": "i"}},
            {"client_uid": {"$regex": q, "$options": "i"}}, {"pan": {"$regex": q, "$options": "i"}},
            {"gstin": {"$regex": q, "$options": "i"}}, {"cin": {"$regex": q, "$options": "i"}},
        ]
    return await list_collection("clients", query)


@api.get("/clients/{client_uid}")
async def get_client(client_uid: str, request: Request):
    await require_user(request)
    c = await db.clients.find_one({"client_uid": client_uid}, {"_id": 0})
    if not c: raise HTTPException(404, "Not found")
    cases = await db.cases.find({"client_uid": client_uid}, {"_id": 0}).to_list(200)
    docs = await db.documents.find({"client_uid": client_uid}, {"_id": 0}).sort("uploaded_at", -1).to_list(200)
    activities = await db.activities.find({"entity_type": {"$in": ["client", "case"]},
                                           "entity_id": {"$in": [client_uid] + [x["case_uid"] for x in cases]}},
                                          {"_id": 0}).sort("created_at", -1).to_list(300)
    invoices = await db.invoices.find({"client_uid": client_uid}, {"_id": 0}).to_list(100)
    payments = await db.payments.find({"client_uid": client_uid}, {"_id": 0}).to_list(100)
    mandates = await db.mandates.find({"client_uid": client_uid}, {"_id": 0}).to_list(50)
    return {"client": c, "cases": cases, "documents": docs, "activities": activities,
            "invoices": invoices, "payments": payments, "mandates": mandates}


@api.patch("/clients/{client_uid}")
async def update_client(client_uid: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    before = await db.clients.find_one({"client_uid": client_uid}, {"_id": 0})
    if not before: raise HTTPException(404, "Not found")
    allowed = {"name", "company", "pan", "cin", "gstin", "mobile", "email", "city", "state",
               "industry", "constitution", "incorporation_date", "relationship_manager", "tags"}
    updates = {k: v for k, v in body.items() if k in allowed}
    if updates:
        await db.clients.update_one({"client_uid": client_uid}, {"$set": updates})
        after = await db.clients.find_one({"client_uid": client_uid}, {"_id": 0})
        await audit(user, "client", client_uid, "updated", before, after)
        return after
    return before


# ============== CASES ==============
@api.get("/cases")
async def get_cases(request: Request, stage: Optional[str] = None, client_uid: Optional[str] = None,
                    product: Optional[str] = None, q: Optional[str] = None):
    user = await require_user(request)
    query = await scope_query(user, "cases")
    if stage: query["stage"] = stage
    if client_uid: query["client_uid"] = client_uid
    if product: query["product"] = product
    if q:
        query["$or"] = [
            {"case_uid": {"$regex": q, "$options": "i"}},
            {"client_uid": {"$regex": q, "$options": "i"}},
            {"purpose": {"$regex": q, "$options": "i"}},
        ]
    rows = await list_collection("cases", query)
    if user.get("role") == "channel_partner":
        rows = [sanitize_partner_case(c) for c in rows]
    return rows


@api.post("/cases")
async def create_case(request: Request):
    user = await require_user(request)
    body = await request.json()
    case_uid = await gen_uid(db, "case")
    now_iso = datetime.now(timezone.utc).isoformat()
    doc = {
        "case_uid": case_uid, "client_uid": body["client_uid"],
        "product": body.get("product", "business_loan"),
        "requirement": float(body.get("requirement", 0)),
        "purpose": body.get("purpose", ""), "tenure_months": int(body.get("tenure_months", 60)),
        "security": body.get("security", ""), "geography": body.get("geography", ""),
        "expected_roi": float(body.get("expected_roi", 11.0)),
        "urgency": body.get("urgency", "normal"),
        "sales_owner": body.get("sales_owner") or user.get("employee_uid"),
        "credit_owner": body.get("credit_owner"),
        "channel_partner_uid": body.get("channel_partner_uid"),
        "source": body.get("source", "manual"), "stage": body.get("stage", "qualified"),
        "expected_closure": body.get("expected_closure"),
        "expected_revenue": float(body.get("requirement", 0)) * 0.012,
        "actual_revenue": 0, "sanctioned_amount": 0, "disbursed_amount": 0,
        "lead_uid": body.get("lead_uid"),
        "documentation_pct": 0, "created_at": now_iso,
    }
    await db.cases.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "case", case_uid, "created", None, doc)
    await push_activity("case", case_uid, "note", user, f"Case created for client {doc['client_uid']}")
    return doc


@api.get("/cases/{case_uid}")
async def get_case(case_uid: str, request: Request):
    await require_user(request)
    c = await db.cases.find_one({"case_uid": case_uid}, {"_id": 0})
    if not c: raise HTTPException(404, "Not found")
    client = await db.clients.find_one({"client_uid": c["client_uid"]}, {"_id": 0})
    apps = await db.applications.find({"case_uid": case_uid}, {"_id": 0}).to_list(100)
    sanctions = await db.sanctions.find({"case_uid": case_uid}, {"_id": 0}).to_list(100)
    disbursements = await db.disbursements.find({"case_uid": case_uid}, {"_id": 0}).to_list(100)
    documents = await db.documents.find({"case_uid": case_uid}, {"_id": 0}).sort("uploaded_at", -1).to_list(200)
    mandate = await db.mandates.find_one({"case_uid": case_uid}, {"_id": 0})
    invoices = await db.invoices.find({"case_uid": case_uid}, {"_id": 0}).to_list(50)
    pds = await db.pds.find({"case_uid": case_uid}, {"_id": 0}).sort("version", -1).to_list(20)
    activities = await db.activities.find({"entity_type": "case", "entity_id": case_uid}, {"_id": 0}).sort("created_at", -1).to_list(200)
    tasks = await db.tasks.find({"case_uid": case_uid}, {"_id": 0}).to_list(100)
    bureau = await db.bureau_checks.find({"case_uid": case_uid}, {"_id": 0}).sort("pulled_at", -1).to_list(20)
    assessment = await db.assessments.find_one({"case_uid": case_uid}, {"_id": 0})
    queries = await db.lender_queries.find({"case_uid": case_uid}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"case": c, "client": client, "applications": apps, "sanctions": sanctions,
            "disbursements": disbursements, "documents": documents, "mandate": mandate,
            "invoices": invoices, "pds": pds, "activities": activities, "tasks": tasks,
            "bureau": bureau, "assessment": assessment, "queries": queries}


@api.patch("/cases/{case_uid}")
async def update_case(case_uid: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    before = await db.cases.find_one({"case_uid": case_uid}, {"_id": 0})
    if not before: raise HTTPException(404, "Not found")
    allowed = {"stage", "product", "requirement", "purpose", "tenure_months", "security",
               "geography", "expected_roi", "urgency", "sales_owner", "credit_owner",
               "channel_partner_uid", "expected_closure", "documentation_pct"}
    updates = {k: v for k, v in body.items() if k in allowed}
    if updates:
        await db.cases.update_one({"case_uid": case_uid}, {"$set": updates})
        after = await db.cases.find_one({"case_uid": case_uid}, {"_id": 0})
        await audit(user, "case", case_uid, "updated", before, after)
        if "stage" in updates and updates["stage"] != before.get("stage"):
            await push_activity("case", case_uid, "status_change", user,
                                f"Stage: {before.get('stage')} → {updates['stage']}")
        return after
    return before


# ============== PD ==============
@api.post("/cases/{case_uid}/pd")
async def create_pd(case_uid: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    c = await db.cases.find_one({"case_uid": case_uid}, {"_id": 0})
    if not c: raise HTTPException(404, "Case not found")
    latest = await db.pds.find_one({"case_uid": case_uid}, {"_id": 0}, sort=[("version", -1)])
    version = (latest["version"] + 1) if latest else 1
    doc = {
        "pd_id": f"pd_{uuid.uuid4().hex[:12]}",
        "case_uid": case_uid, "client_uid": c["client_uid"],
        "template": body.get("template", "business"),
        "data": body.get("data", {}),
        "version": version, "conducted_by": user["user_id"],
        "conducted_on": datetime.now(timezone.utc).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.pds.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "case", case_uid, "pd_created", None, {"pd_id": doc["pd_id"], "version": version})
    await push_activity("case", case_uid, "note", user, f"PD v{version} recorded")
    return doc


# ============== DOCUMENTS ==============
@api.get("/documents")
async def list_documents(request: Request, case_uid: Optional[str] = None, client_uid: Optional[str] = None,
                         category: Optional[str] = None):
    await require_user(request)
    query = {}
    if case_uid: query["case_uid"] = case_uid
    if client_uid: query["client_uid"] = client_uid
    if category: query["category"] = category
    return await list_collection("documents", query, sort_field="uploaded_at")


@api.post("/documents/upload")
async def upload_document(request: Request, file: UploadFile = File(...),
                          client_uid: str = Form(...), case_uid: Optional[str] = Form(None),
                          category: str = Form("Other"), doc_type: str = Form(""),
                          financial_period: Optional[str] = Form(None),
                          notes: Optional[str] = Form("")):
    user = await require_user(request)
    ext = file.filename.split(".")[-1] if "." in file.filename else "bin"
    doc_id = f"doc_{uuid.uuid4().hex[:12]}"
    path = f"{APP_NAME}/documents/{client_uid}/{doc_id}.{ext}"
    data = await file.read()
    try:
        result = put_object(path, data, file.content_type or "application/octet-stream")
        storage_path = result.get("path", path)
        size = result.get("size", len(data))
    except Exception as e:
        log.error(f"Upload failed: {e}")
        raise HTTPException(500, "Upload failed")

    latest = await db.documents.find_one({"client_uid": client_uid, "case_uid": case_uid, "doc_type": doc_type},
                                          {"_id": 0}, sort=[("version", -1)])
    version = (latest["version"] + 1) if latest else 1
    if latest:
        await db.documents.update_one({"document_id": latest["document_id"]}, {"$set": {"superseded_by": doc_id}})

    doc = {
        "document_id": doc_id, "client_uid": client_uid, "case_uid": case_uid,
        "category": category, "doc_type": doc_type, "financial_period": financial_period,
        "version": version, "storage_path": storage_path,
        "original_filename": file.filename, "content_type": file.content_type,
        "size": size, "uploaded_by": user["user_id"],
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "status": "received", "verified_by": None, "verified_at": None,
        "expiry_date": None, "tags": [], "notes": notes, "superseded_by": None,
    }
    await db.documents.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "document", doc_id, "uploaded", None, {"filename": file.filename, "case_uid": case_uid})
    if case_uid:
        await push_activity("case", case_uid, "upload", user, f"Document uploaded: {file.filename}")
        # recompute doc completeness: 12 default categories, credit 1 per unique category present
        cats = await db.documents.distinct("category", {"case_uid": case_uid})
        pct = min(100, int(round(len(cats) * 100 / 12)))
        await db.cases.update_one({"case_uid": case_uid}, {"$set": {"documentation_pct": pct}})
    return doc


@api.get("/documents/{document_id}/download")
async def download_document(document_id: str, request: Request):
    await require_user(request)
    doc = await db.documents.find_one({"document_id": document_id}, {"_id": 0})
    if not doc or not doc.get("storage_path"):
        raise HTTPException(404, "Not found")
    try:
        data, ct = get_object(doc["storage_path"])
    except Exception as e:
        raise HTTPException(500, f"Download failed: {e}")
    return FastResponse(content=data, media_type=doc.get("content_type") or ct,
                        headers={"Content-Disposition": f'attachment; filename=\"{doc["original_filename"]}\"'})


@api.patch("/documents/{document_id}")
async def update_document(document_id: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    before = await db.documents.find_one({"document_id": document_id}, {"_id": 0})
    if not before: raise HTTPException(404, "Not found")
    allowed = {"status", "verified_by", "verified_at", "expiry_date", "tags", "notes", "category", "doc_type"}
    updates = {k: v for k, v in body.items() if k in allowed}
    if "status" in updates and updates["status"] == "valid":
        updates.setdefault("verified_by", user["user_id"])
        updates.setdefault("verified_at", datetime.now(timezone.utc).isoformat())
    await db.documents.update_one({"document_id": document_id}, {"$set": updates})
    after = await db.documents.find_one({"document_id": document_id}, {"_id": 0})
    await audit(user, "document", document_id, "updated", before, after)
    return after


# ============== BUREAU (Sandbox) ==============
@api.post("/cases/{case_uid}/bureau")
async def pull_bureau(case_uid: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    c = await db.cases.find_one({"case_uid": case_uid}, {"_id": 0})
    if not c: raise HTTPException(404, "Case not found")
    if not body.get("consent"): raise HTTPException(400, "Consent required")
    provider = body.get("provider", "cibil")

    import random as _r
    doc = {
        "bureau_id": f"br_{uuid.uuid4().hex[:12]}",
        "client_uid": c["client_uid"], "case_uid": case_uid,
        "provider": provider, "consent_captured": True,
        "consent_at": datetime.now(timezone.utc).isoformat(),
        "pulled_at": datetime.now(timezone.utc).isoformat(),
        "reference_number": f"SANDBOX-{provider.upper()}-{uuid.uuid4().hex[:8]}",
        "score": _r.randint(650, 810),
        "accounts": _r.randint(3, 12), "enquiries": _r.randint(1, 8),
        "dpd_current": _r.randint(0, 15), "overdue_amount": 0,
        "written_off": 0, "settlements": 0,
        "utilisation": round(_r.uniform(20, 75), 1),
        "raw_report": {"note": "Sandbox report - not a real bureau pull"},
        "is_sandbox": True,
    }
    await db.bureau_checks.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "case", case_uid, "bureau_pulled", None, {"provider": provider, "score": doc["score"], "sandbox": True})
    await push_activity("case", case_uid, "note", user, f"Bureau pulled ({provider.upper()}) — Score {doc['score']} [SANDBOX]")
    return doc


# ============== CREDIT ASSESSMENT ==============
@api.post("/cases/{case_uid}/assessment")
async def save_assessment(case_uid: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    c = await db.cases.find_one({"case_uid": case_uid}, {"_id": 0})
    if not c: raise HTTPException(404, "Case not found")

    # Validate flags shape — reject anything that isn't a dict with allowed level + non-empty title.
    raw_flags = body.get("flags", []) or []
    if not isinstance(raw_flags, list):
        raise HTTPException(422, "flags must be a list")
    cleaned_flags = []
    for i, fl in enumerate(raw_flags):
        if not isinstance(fl, dict):
            raise HTTPException(422, f"flags[{i}] must be an object with level & title")
        lvl = str(fl.get("level", "")).lower()
        title = str(fl.get("title", "")).strip()
        if lvl not in ("green", "amber", "red"):
            raise HTTPException(422, f"flags[{i}].level must be green|amber|red")
        if not title:
            raise HTTPException(422, f"flags[{i}].title required")
        cleaned_flags.append({
            "level": lvl,
            "title": title[:240],
            "id": fl.get("id") or f"fl_{uuid.uuid4().hex[:8]}",
        })

    doc = {
        "assessment_id": f"ass_{uuid.uuid4().hex[:12]}",
        "case_uid": case_uid, "client_uid": c["client_uid"],
        "overview": body.get("overview", {}), "financials": body.get("financials", {}),
        "banking": body.get("banking", {}), "ratios": body.get("ratios", {}),
        "positives": body.get("positives", []), "concerns": body.get("concerns", []),
        "flags": cleaned_flags,
        "indicative_eligibility": body.get("indicative_eligibility", 0),
        "analyst_comments": body.get("analyst_comments", ""),
        "recommendation": body.get("recommendation", ""),
        "prepared_by": user["user_id"],
        "prepared_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.assessments.replace_one({"case_uid": case_uid}, doc, upsert=True)
    await audit(user, "case", case_uid, "assessment_saved", None, {"assessment_id": doc["assessment_id"]})
    return doc


# ============== MANDATES ==============
@api.get("/mandates")
async def list_mandates(request: Request):
    await require_user(request)
    return await list_collection("mandates")


@api.post("/mandates")
async def create_mandate(request: Request):
    user = await require_user(request)
    body = await request.json()
    m_uid = await gen_uid(db, "mandate")
    doc = {
        "mandate_uid": m_uid, "client_uid": body["client_uid"], "case_uid": body["case_uid"],
        "scope": body.get("scope", ""),
        "upfront_fee": float(body.get("upfront_fee", 0)),
        "success_fee_pct": float(body.get("success_fee_pct", 1.0)),
        "min_fee": float(body.get("min_fee", 0)),
        "taxes_pct": float(body.get("taxes_pct", 18)),
        "other_charges": float(body.get("other_charges", 0)),
        "validity_days": int(body.get("validity_days", 90)),
        "exclusivity": bool(body.get("exclusivity", False)),
        "signatory": body.get("signatory", ""), "signing_method": body.get("signing_method", "esign"),
        "status": "draft", "version": 1, "signed_at": None, "document_id": None,
        "created_by": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.mandates.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "mandate", m_uid, "created", None, doc)
    return doc


@api.patch("/mandates/{mandate_uid}")
async def update_mandate(mandate_uid: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    before = await db.mandates.find_one({"mandate_uid": mandate_uid}, {"_id": 0})
    if not before: raise HTTPException(404, "Not found")
    allowed = {"status", "scope", "upfront_fee", "success_fee_pct", "min_fee", "taxes_pct",
               "other_charges", "validity_days", "exclusivity", "signatory", "signing_method", "signed_at"}
    updates = {k: v for k, v in body.items() if k in allowed}
    if updates.get("status") == "signed" and not updates.get("signed_at"):
        updates["signed_at"] = datetime.now(timezone.utc).isoformat()
    await db.mandates.update_one({"mandate_uid": mandate_uid}, {"$set": updates})
    after = await db.mandates.find_one({"mandate_uid": mandate_uid}, {"_id": 0})
    await audit(user, "mandate", mandate_uid, "updated", before, after)
    return after


# ============== INVOICES & PAYMENTS ==============
@api.get("/invoices")
async def list_invoices(request: Request):
    await require_user(request)
    return await list_collection("invoices")


@api.post("/invoices")
async def create_invoice(request: Request):
    user = await require_user(request)
    body = await request.json()
    if body.get("amount") is None or not body.get("client_uid"):
        raise HTTPException(422, "amount and client_uid required")
    uid = await gen_uid(db, "invoice")
    amt = float(body["amount"])
    gst = float(body.get("gst_pct", 18))
    tds = float(body.get("tds_pct", 0))
    doc = {
        "invoice_uid": uid, "client_uid": body["client_uid"],
        "case_uid": body.get("case_uid"), "mandate_uid": body.get("mandate_uid"),
        "amount": amt, "gst_pct": gst, "tds_pct": tds,
        "total_amount": round(amt * (1 + gst / 100) * (1 - tds / 100), 2),
        "status": "pending", "due_date": body.get("due_date"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.invoices.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "invoice", uid, "created", None, doc)
    return doc


@api.get("/payments")
async def list_payments(request: Request):
    await require_user(request)
    return await list_collection("payments")


@api.post("/payments")
async def create_payment(request: Request):
    user = await require_user(request)
    body = await request.json()
    if body.get("amount") is None or not body.get("invoice_uid") or not body.get("client_uid"):
        raise HTTPException(422, "invoice_uid, client_uid and amount required")
    uid = await gen_uid(db, "payment")
    doc = {
        "payment_uid": uid, "invoice_uid": body["invoice_uid"],
        "client_uid": body["client_uid"], "amount": float(body["amount"]),
        "mode": body.get("mode", "neft"), "reference": body.get("reference", ""),
        "received_on": body.get("received_on", datetime.now(timezone.utc).isoformat()),
        "remarks": body.get("remarks", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.payments.insert_one(doc)
    doc.pop("_id", None)
    # update invoice status using total_amount (GST/TDS aware)
    inv = await db.invoices.find_one({"invoice_uid": body["invoice_uid"]}, {"_id": 0})
    if inv:
        payments = await db.payments.find({"invoice_uid": inv["invoice_uid"]}, {"_id": 0}).to_list(50)
        total_paid = sum(p.get("amount", 0) for p in payments)
        target = inv.get("total_amount") or inv.get("amount", 0)
        new_status = "paid" if total_paid + 0.01 >= target else "part_paid"
        await db.invoices.update_one({"invoice_uid": inv["invoice_uid"]}, {"$set": {"status": new_status}})
    await audit(user, "payment", uid, "created", None, doc)
    return doc


# ============== LENDERS ==============
@api.get("/lenders")
async def list_lenders(request: Request):
    await require_user(request)
    return await list_collection("lenders", sort_field="name")


@api.post("/lenders")
async def create_lender(request: Request):
    user = await require_user(request)
    body = await request.json()
    body["lender_id"] = f"lender_{uuid.uuid4().hex[:10]}"
    body.setdefault("created_at", datetime.now(timezone.utc).isoformat())
    body.setdefault("active", True)
    await db.lenders.insert_one(body)
    body.pop("_id", None)
    await audit(user, "lender", body["lender_id"], "created", None, body)
    return body


@api.patch("/lenders/{lender_id}")
async def update_lender(lender_id: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    before = await db.lenders.find_one({"lender_id": lender_id}, {"_id": 0})
    if not before: raise HTTPException(404, "Not found")
    await db.lenders.update_one({"lender_id": lender_id}, {"$set": body})
    after = await db.lenders.find_one({"lender_id": lender_id}, {"_id": 0})
    await audit(user, "lender", lender_id, "updated", before, after)
    return after


@api.get("/lenders/suggest/{case_uid}")
async def suggest_lenders(case_uid: str, request: Request):
    await require_user(request)
    c = await db.cases.find_one({"case_uid": case_uid}, {"_id": 0})
    if not c: raise HTTPException(404, "Case not found")
    lenders = await db.lenders.find({"active": True}, {"_id": 0}).to_list(200)
    scored = []
    for l in lenders:
        score = 100
        if c["product"] not in l.get("products", []): score -= 40
        if c["requirement"] < l.get("ticket_size_min", 0): score -= 20
        if c["requirement"] > l.get("ticket_size_max", 0): score -= 20
        scored.append({"lender": l, "score": score})
    scored.sort(key=lambda x: -x["score"])
    return scored[:10]


# ============== LENDER APPLICATIONS ==============
@api.get("/applications")
async def list_applications(request: Request, status: Optional[str] = None, case_uid: Optional[str] = None):
    await require_user(request)
    q = {}
    if status: q["status"] = status
    if case_uid: q["case_uid"] = case_uid
    return await list_collection("applications", q)


@api.post("/applications")
async def create_application(request: Request):
    user = await require_user(request)
    body = await request.json()
    uid = await gen_uid(db, "application")
    c = await db.cases.find_one({"case_uid": body["case_uid"]}, {"_id": 0})
    if not c: raise HTTPException(404, "Case not found")
    doc = {
        "application_uid": uid, "case_uid": body["case_uid"], "client_uid": c["client_uid"],
        "lender_id": body["lender_id"], "submission_date": body.get("submission_date"),
        "lender_rm": body.get("lender_rm", ""), "lender_login_no": body.get("lender_login_no", ""),
        "amount_requested": float(body.get("amount_requested", c["requirement"])),
        "product": body.get("product", c["product"]),
        "status": body.get("status", "mapped"),
        "sanction_amount": 0, "roi": 0, "tenure_months": 0, "fees": 0,
        "security": body.get("security", c.get("security", "")),
        "conditions": "", "rejection_reason": "",
        "documents_shared": body.get("documents_shared", []),
        "created_by": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.applications.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "application", uid, "created", None, doc)
    await push_activity("case", body["case_uid"], "note", user, f"Lender application created ({body['lender_id']})")
    return doc


@api.patch("/applications/{application_uid}")
async def update_application(application_uid: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    before = await db.applications.find_one({"application_uid": application_uid}, {"_id": 0})
    if not before: raise HTTPException(404, "Not found")
    allowed = {"status", "submission_date", "lender_rm", "lender_login_no", "amount_requested",
               "sanction_amount", "roi", "tenure_months", "fees", "security", "conditions", "rejection_reason"}
    updates = {k: v for k, v in body.items() if k in allowed}
    await db.applications.update_one({"application_uid": application_uid}, {"$set": updates})
    after = await db.applications.find_one({"application_uid": application_uid}, {"_id": 0})
    await audit(user, "application", application_uid, "updated", before, after)
    await push_activity("case", before["case_uid"], "status_change", user,
                        f"Application {application_uid}: {before.get('status')} → {after.get('status')}")
    return after


# ============== LENDER QUERIES ==============
@api.get("/lender-queries")
async def list_queries(request: Request, application_uid: Optional[str] = None, case_uid: Optional[str] = None):
    await require_user(request)
    q = {}
    if application_uid: q["application_uid"] = application_uid
    if case_uid: q["case_uid"] = case_uid
    return await list_collection("lender_queries", q)


@api.post("/lender-queries")
async def create_query(request: Request):
    user = await require_user(request)
    body = await request.json()
    doc = {
        "query_id": f"q_{uuid.uuid4().hex[:10]}",
        "application_uid": body["application_uid"], "case_uid": body["case_uid"],
        "query_text": body["query_text"], "raised_by": body.get("raised_by", "lender"),
        "assigned_to": body.get("assigned_to"),
        "required_document": body.get("required_document"),
        "due_date": body.get("due_date"),
        "response": None, "attachment_id": None,
        "status": "open",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.lender_queries.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "lender_query", doc["query_id"], "created", None, doc)
    return doc


@api.patch("/lender-queries/{query_id}")
async def update_query(query_id: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    before = await db.lender_queries.find_one({"query_id": query_id}, {"_id": 0})
    if not before: raise HTTPException(404, "Not found")
    await db.lender_queries.update_one({"query_id": query_id}, {"$set": body})
    after = await db.lender_queries.find_one({"query_id": query_id}, {"_id": 0})
    await audit(user, "lender_query", query_id, "updated", before, after)
    return after


# ============== SANCTIONS ==============
@api.get("/sanctions")
async def list_sanctions(request: Request):
    await require_user(request)
    return await list_collection("sanctions", sort_field="sanction_date")


@api.post("/sanctions")
async def create_sanction(request: Request):
    user = await require_user(request)
    body = await request.json()
    uid = await gen_uid(db, "sanction")
    doc = {
        "sanction_uid": uid, "application_uid": body["application_uid"],
        "case_uid": body["case_uid"], "lender_id": body["lender_id"],
        "sanction_amount": float(body["sanction_amount"]),
        "sanction_date": body.get("sanction_date", datetime.now(timezone.utc).isoformat()),
        "roi": float(body.get("roi", 0)), "benchmark": body.get("benchmark", "REPO"),
        "spread": float(body.get("spread", 0)),
        "tenure_months": int(body.get("tenure_months", 60)),
        "emi": float(body.get("emi", 0)), "moratorium_months": int(body.get("moratorium_months", 0)),
        "security": body.get("security", ""), "ltv": float(body.get("ltv", 0)),
        "processing_fee_pct": float(body.get("processing_fee_pct", 0)),
        "insurance_amount": float(body.get("insurance_amount", 0)),
        "conditions_precedent": body.get("conditions_precedent", ""),
        "conditions_subsequent": body.get("conditions_subsequent", ""),
        "validity_days": int(body.get("validity_days", 30)),
        "document_id": body.get("document_id"),
        "status": body.get("status", "received"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.sanctions.insert_one(doc)
    doc.pop("_id", None)
    # update case
    await db.cases.update_one({"case_uid": body["case_uid"]},
                              {"$inc": {"sanctioned_amount": doc["sanction_amount"]},
                               "$set": {"stage": "sanctioned"}})
    await db.applications.update_one({"application_uid": body["application_uid"]},
                                     {"$set": {"status": "sanctioned",
                                               "sanction_amount": doc["sanction_amount"],
                                               "roi": doc["roi"], "tenure_months": doc["tenure_months"]}})
    await audit(user, "sanction", uid, "created", None, doc)
    await push_activity("case", body["case_uid"], "status_change", user,
                        f"Sanction received: ₹{doc['sanction_amount']:,.0f} @ {doc['roi']}%")
    return doc


# ============== DISBURSEMENTS ==============
@api.get("/disbursements")
async def list_disbursements(request: Request):
    await require_user(request)
    return await list_collection("disbursements", sort_field="disbursement_date")


@api.post("/disbursements")
async def create_disbursement(request: Request):
    user = await require_user(request)
    body = await request.json()
    uid = await gen_uid(db, "disbursement")
    case_uid = body["case_uid"]
    s = await db.sanctions.find_one({"sanction_uid": body["sanction_uid"]}, {"_id": 0})
    if not s: raise HTTPException(404, "Sanction not found")

    tranche_count = await db.disbursements.count_documents({"sanction_uid": body["sanction_uid"]})
    doc = {
        "disbursement_uid": uid, "case_uid": case_uid, "sanction_uid": body["sanction_uid"],
        "lender_id": s["lender_id"], "amount": float(body["amount"]),
        "requested_date": body.get("requested_date"),
        "disbursement_date": body.get("disbursement_date", datetime.now(timezone.utc).isoformat()),
        "reference": body.get("reference", ""), "destination": body.get("destination", ""),
        "status": "completed", "notes": body.get("notes", ""),
        "proof_document_id": body.get("proof_document_id"), "tranche_no": tranche_count + 1,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.disbursements.insert_one(doc)
    doc.pop("_id", None)
    # Update case
    await db.cases.update_one({"case_uid": case_uid}, {"$inc": {"disbursed_amount": doc["amount"]}})
    c = await db.cases.find_one({"case_uid": case_uid}, {"_id": 0})
    new_stage = "fully_disbursed" if c["disbursed_amount"] >= c["sanctioned_amount"] else "partially_disbursed"
    await db.cases.update_one({"case_uid": case_uid}, {"$set": {"stage": new_stage, "actual_revenue": c["expected_revenue"]}})
    await audit(user, "disbursement", uid, "created", None, doc)
    await push_activity("case", case_uid, "status_change", user,
                        f"Disbursement: ₹{doc['amount']:,.0f} (tranche {doc['tranche_no']})")

    # Auto-create incentive & CP commission accrual
    if c.get("sales_owner"):
        agent_uid = c["sales_owner"]
        period = datetime.now(timezone.utc).strftime("%Y-%m")
        await db.incentives.insert_one({
            "incentive_id": f"inc_{uid[-6:]}",
            "employee_uid": agent_uid, "period": period,
            "disbursement_amount": doc["amount"], "revenue_collected": doc["amount"] * 0.01,
            "calculated_amount": doc["amount"] * 0.001,
            "override_amount": None, "override_reason": None, "status": "accrued",
            "approved_by": None, "paid_on": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    if c.get("channel_partner_uid"):
        await db.cp_commissions.insert_one({
            "commission_id": f"cpc_{uid[-6:]}",
            "partner_uid": c["channel_partner_uid"], "case_uid": case_uid,
            "disbursement_uid": uid, "disbursement_amount": doc["amount"],
            "commission_pct": 1.0, "commission_amount": doc["amount"] * 0.01,
            "tds_amount": doc["amount"] * 0.001, "payable_amount": doc["amount"] * 0.009,
            "status": "accrued", "paid_on": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    return doc


# ============== EMPLOYEES ==============
@api.get("/employees")
async def list_employees(request: Request):
    await require_user(request)
    return await list_collection("employees", sort_field="employee_uid")


@api.patch("/employees/{employee_uid}")
async def update_employee(employee_uid: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    before = await db.employees.find_one({"employee_uid": employee_uid}, {"_id": 0})
    if not before: raise HTTPException(404, "Not found")
    await db.employees.update_one({"employee_uid": employee_uid}, {"$set": body})
    after = await db.employees.find_one({"employee_uid": employee_uid}, {"_id": 0})
    if "role" in body:
        await db.users.update_one({"employee_uid": employee_uid}, {"$set": {"role": body["role"]}})
    await audit(user, "employee", employee_uid, "updated", before, after)
    return after


# ============== INCENTIVES ==============
@api.get("/incentives")
async def list_incentives(request: Request, employee_uid: Optional[str] = None):
    await require_user(request)
    q = {}
    if employee_uid: q["employee_uid"] = employee_uid
    return await list_collection("incentives")


@api.patch("/incentives/{incentive_id}")
async def update_incentive(incentive_id: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    before = await db.incentives.find_one({"incentive_id": incentive_id}, {"_id": 0})
    if not before: raise HTTPException(404, "Not found")
    await db.incentives.update_one({"incentive_id": incentive_id}, {"$set": body})
    after = await db.incentives.find_one({"incentive_id": incentive_id}, {"_id": 0})
    await audit(user, "incentive", incentive_id, "updated", before, after)
    return after


# ============== CHANNEL PARTNERS ==============
@api.get("/channel-partners")
async def list_cps(request: Request):
    await require_user(request)
    return await list_collection("channel_partners", sort_field="name")


@api.post("/channel-partners")
async def create_cp(request: Request):
    user = await require_user(request)
    if not has_permission(user, "create_partner"):
        raise HTTPException(403, "Missing permission: create_partner")
    body = await request.json()
    uid = await gen_uid(db, "channel_partner")
    code = f"CP-{(body.get('city') or 'IND')[:3].upper()}-{uid.split('-')[-1]}"
    doc = {
        "partner_uid": uid, "channel_code": code, "name": body["name"],
        "entity_name": body.get("entity_name", ""), "pan": body.get("pan", ""),
        "gst": body.get("gst", ""), "kyc_status": "pending",
        "bank_account": body.get("bank_account", {}), "agreement_signed": False,
        "products": body.get("products", []), "geography": body.get("geography", []),
        "channel_manager_uid": body.get("channel_manager_uid"),
        "commission_structure": body.get("commission_structure", {"default_pct": 1.0}),
        "status": "active",
        "mobile": body.get("mobile", ""), "email": body.get("email", ""),
        "city": body.get("city", ""), "state": body.get("state", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.channel_partners.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "channel_partner", uid, "created", None, doc)
    return doc


@api.get("/cp-commissions")
async def list_cp_commissions(request: Request, partner_uid: Optional[str] = None):
    await require_user(request)
    q = {}
    if partner_uid: q["partner_uid"] = partner_uid
    return await list_collection("cp_commissions", q)


# ============== TASKS ==============
@api.get("/tasks")
async def list_tasks(request: Request, owner_uid: Optional[str] = None, status: Optional[str] = None):
    user = await require_user(request)
    q = await scope_query(user, "tasks")
    if owner_uid: q["owner_uid"] = owner_uid
    if status: q["status"] = status
    return await list_collection("tasks", q, sort_field="due_date")


@api.post("/tasks")
async def create_task(request: Request):
    user = await require_user(request)
    body = await request.json()
    if not body.get("title"):
        raise HTTPException(422, "title required")
    uid = await gen_uid(db, "task")
    doc = {
        "task_id": uid, "title": body["title"],
        "description": body.get("description", ""),
        "owner_uid": body.get("owner_uid") or user.get("employee_uid"),
        "created_by": user["user_id"],
        "case_uid": body.get("case_uid"), "client_uid": body.get("client_uid"),
        "lead_uid": body.get("lead_uid"), "application_uid": body.get("application_uid"),
        "priority": body.get("priority", "normal"),
        "due_date": body.get("due_date") or (datetime.now(timezone.utc) + timedelta(days=3)).isoformat(),
        "status": "open",
        "origin": body.get("origin", "manual"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.tasks.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "task", uid, "created", None, doc)
    return doc


@api.patch("/tasks/{task_id}")
async def update_task(task_id: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    before = await db.tasks.find_one({"task_id": task_id}, {"_id": 0})
    if not before: raise HTTPException(404, "Not found")
    await db.tasks.update_one({"task_id": task_id}, {"$set": body})
    after = await db.tasks.find_one({"task_id": task_id}, {"_id": 0})
    await audit(user, "task", task_id, "updated", before, after)
    return after


# ============== AUDIT LOG ==============
@api.get("/audit-logs")
async def list_audit(request: Request, entity_type: Optional[str] = None, entity_id: Optional[str] = None):
    await require_user(request)
    q = {}
    if entity_type: q["entity_type"] = entity_type
    if entity_id: q["entity_id"] = entity_id
    return await list_collection("audit_logs", q, sort_field="at", limit=500)


# ============== USERS / ROLES ADMIN ==============
@api.get("/users")
async def list_users(request: Request):
    user = await require_user(request)
    if not has_permission(user, "manage_users"):
        raise HTTPException(403, "Missing permission: manage_users")
    users = await db.users.find({}, {"_id": 0}).to_list(500)
    return users


@api.patch("/users/{user_id}")
async def update_user(user_id: str, request: Request):
    actor = await require_user(request)
    if actor["role"] != "super_admin":
        raise HTTPException(403, "Admin only")
    body = await request.json()
    before = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not before: raise HTTPException(404, "Not found")
    allowed = {"role", "active", "name"}
    updates = {k: v for k, v in body.items() if k in allowed}
    if updates:
        await db.users.update_one({"user_id": user_id}, {"$set": updates})
        if "role" in updates and before.get("employee_uid"):
            await db.employees.update_one({"employee_uid": before["employee_uid"]}, {"$set": {"role": updates["role"]}})
        after = await db.users.find_one({"user_id": user_id}, {"_id": 0})
        await audit(actor, "user", user_id, "updated", before, after)
        return after
    return before


# ============== INTEGRATIONS STATUS ==============
@api.get("/integrations")
async def list_integrations(request: Request):
    await require_user(request)
    return [
        {"key": "google_drive", "name": "Google Drive", "status": "not_connected", "category": "data"},
        {"key": "google_sheets", "name": "Google Sheets", "status": "not_connected", "category": "data"},
        {"key": "bureau_cibil", "name": "CIBIL", "status": "sandbox", "category": "bureau"},
        {"key": "bureau_experian", "name": "Experian", "status": "sandbox", "category": "bureau"},
        {"key": "bureau_crif", "name": "CRIF High Mark", "status": "sandbox", "category": "bureau"},
        {"key": "bureau_equifax", "name": "Equifax", "status": "sandbox", "category": "bureau"},
        {"key": "esign", "name": "eSign", "status": "sandbox", "category": "signing"},
        {"key": "payment_gateway", "name": "Payment Gateway", "status": "sandbox", "category": "payment"},
        {"key": "email", "name": "Email", "status": "not_connected", "category": "comms"},
        {"key": "whatsapp", "name": "WhatsApp", "status": "not_connected", "category": "comms"},
        {"key": "telephony", "name": "Telephony", "status": "not_connected", "category": "comms"},
        {"key": "sms", "name": "SMS", "status": "not_connected", "category": "comms"},
        {"key": "google_auth", "name": "Google Auth (Emergent)", "status": "connected", "category": "auth"},
        {"key": "object_storage", "name": "Object Storage (Emergent)", "status": "connected", "category": "storage"},
    ]


# ============== GLOBAL SEARCH ==============
@api.get("/search")
async def global_search(request: Request, q: str = Query(..., min_length=1)):
    await require_user(request)
    # escape regex special chars so PAN/GSTIN with dots etc match literally
    q_esc = re.escape(q.strip())
    regex = {"$regex": q_esc, "$options": "i"}
    leads = await db.leads.find({"$or": [{"lead_uid": regex}, {"name": regex}, {"company": regex},
                                          {"mobile": regex}, {"email": regex}, {"pan": regex}, {"gstin": regex}]},
                                {"_id": 0}).limit(8).to_list(8)
    clients = await db.clients.find({"$or": [{"client_uid": regex}, {"name": regex}, {"company": regex},
                                              {"mobile": regex}, {"email": regex}, {"pan": regex},
                                              {"gstin": regex}, {"cin": regex}]}, {"_id": 0}).limit(8).to_list(8)
    # cases: by uid AND by owning-client PAN/GSTIN/name via lookup
    case_or = [{"case_uid": regex}, {"purpose": regex}]
    matched_clients = await db.clients.find({"$or": [{"pan": regex}, {"gstin": regex}, {"name": regex}, {"company": regex}]}, {"client_uid": 1}).limit(20).to_list(20)
    if matched_clients:
        case_or.append({"client_uid": {"$in": [c["client_uid"] for c in matched_clients]}})
    cases = await db.cases.find({"$or": case_or}, {"_id": 0}).limit(8).to_list(8)
    apps = await db.applications.find({"$or": [{"application_uid": regex}, {"lender_login_no": regex}, {"case_uid": regex}]}, {"_id": 0}).limit(6).to_list(6)
    sanctions = await db.sanctions.find({"$or": [{"sanction_uid": regex}, {"case_uid": regex}]}, {"_id": 0}).limit(5).to_list(5)
    disbursements = await db.disbursements.find({"$or": [{"disbursement_uid": regex}, {"case_uid": regex}]}, {"_id": 0}).limit(5).to_list(5)
    invoices = await db.invoices.find({"$or": [{"invoice_uid": regex}, {"case_uid": regex}, {"client_uid": regex}]}, {"_id": 0}).limit(5).to_list(5)
    tasks = await db.tasks.find({"$or": [{"task_uid": regex}, {"title": regex}]}, {"_id": 0}).limit(5).to_list(5)
    partners = await db.channel_partners.find({"$or": [{"channel_code": regex}, {"name": regex}, {"partner_uid": regex}, {"contact_mobile": regex}, {"contact_email": regex}]}, {"_id": 0}).limit(6).to_list(6)
    return {
        "leads": leads, "clients": clients, "cases": cases,
        "applications": apps, "sanctions": sanctions, "disbursements": disbursements,
        "invoices": invoices, "tasks": tasks, "partners": partners,
    }


# ============== REPORTS ==============
@api.get("/reports/daily")
async def daily_report(request: Request):
    await require_user(request)
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    employees = await db.employees.find({}, {"_id": 0}).to_list(200)
    result = []
    for e in employees:
        leads = await db.leads.count_documents({"assigned_to": e["employee_uid"], "created_at": {"$gte": today_start}})
        cases = await db.cases.count_documents({"sales_owner": e["employee_uid"], "created_at": {"$gte": today_start}})
        activities = await db.activities.count_documents({"author_uid": e.get("user_id"), "created_at": {"$gte": today_start}})
        result.append({
            "employee_uid": e["employee_uid"], "name": e["name"], "role": e["role"],
            "leads_assigned": leads, "cases_created": cases, "activities": activities,
        })
    return result


@api.get("/reports/pipeline")
async def pipeline_report(request: Request):
    await require_user(request)
    rows = await db.cases.aggregate([
        {"$group": {"_id": "$stage", "count": {"$sum": 1},
                    "requested": {"$sum": "$requirement"},
                    "sanctioned": {"$sum": "$sanctioned_amount"},
                    "disbursed": {"$sum": "$disbursed_amount"}}},
        {"$sort": {"count": -1}},
    ]).to_list(50)
    return [{"stage": r["_id"], "count": r["count"], "requested": r["requested"],
             "sanctioned": r["sanctioned"], "disbursed": r["disbursed"]} for r in rows]


# ============== PUBLIC / CUSTOMER (Urban Money-style) ==============
PUBLIC_PRODUCTS = [
    {"slug": "home-loan", "key": "home_loan", "title": "Home Loan", "tagline": "Own your home from ₹5 L to ₹10 Cr",
     "rate_from": 8.35, "tenure_max": 360, "min_amount": 500000, "max_amount": 100000000,
     "hero_image": "https://images.unsplash.com/photo-1568605114967-8130f3a36994?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
    {"slug": "business-loan", "key": "business_loan", "title": "Business Loan", "tagline": "Unsecured funds to fuel growth",
     "rate_from": 12.0, "tenure_max": 60, "min_amount": 200000, "max_amount": 50000000,
     "hero_image": "https://images.unsplash.com/photo-1454165804606-c3d57bc86b40?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
    {"slug": "lap", "key": "lap", "title": "Loan Against Property", "tagline": "Unlock value from your property",
     "rate_from": 9.25, "tenure_max": 180, "min_amount": 1000000, "max_amount": 250000000,
     "hero_image": "https://images.unsplash.com/photo-1600585154340-be6161a56a0c?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
    {"slug": "personal-loan", "key": "personal_loan", "title": "Personal Loan", "tagline": "Instant unsecured funds up to ₹40 L",
     "rate_from": 10.5, "tenure_max": 60, "min_amount": 50000, "max_amount": 4000000,
     "hero_image": "https://images.unsplash.com/photo-1554224155-6726b3ff858f?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
    {"slug": "working-capital", "key": "working_capital", "title": "Working Capital", "tagline": "OD / CC to keep operations liquid",
     "rate_from": 9.5, "tenure_max": 12, "min_amount": 500000, "max_amount": 500000000,
     "hero_image": "https://images.unsplash.com/photo-1450101499163-c8848c66ca85?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
    {"slug": "cc-od", "key": "cc_od", "title": "CC / OD", "tagline": "Revolving credit for business needs",
     "rate_from": 9.75, "tenure_max": 12, "min_amount": 500000, "max_amount": 500000000,
     "hero_image": "https://images.unsplash.com/photo-1526628953301-3e589a6a8b74?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
    {"slug": "term-loan", "key": "term_loan", "title": "Term Loan", "tagline": "Structured funding for CAPEX & expansion",
     "rate_from": 9.0, "tenure_max": 120, "min_amount": 1000000, "max_amount": 1000000000,
     "hero_image": "https://images.unsplash.com/photo-1507679799987-c73779587ccf?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
    {"slug": "equipment-finance", "key": "equipment_finance", "title": "Equipment Finance", "tagline": "Fund plant & machinery to 90% LTV",
     "rate_from": 10.5, "tenure_max": 84, "min_amount": 500000, "max_amount": 500000000,
     "hero_image": "https://images.unsplash.com/photo-1504917595217-d4dc5ebe6122?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
    {"slug": "project-finance", "key": "project_finance", "title": "Project Finance", "tagline": "Long-tenor debt for greenfield & brownfield",
     "rate_from": 10.0, "tenure_max": 240, "min_amount": 10000000, "max_amount": 5000000000,
     "hero_image": "https://images.unsplash.com/photo-1541888946425-d81bb19240f5?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
    {"slug": "construction-finance", "key": "construction_finance", "title": "Construction Finance", "tagline": "Draw-based funding for developers",
     "rate_from": 12.5, "tenure_max": 60, "min_amount": 20000000, "max_amount": 5000000000,
     "hero_image": "https://images.unsplash.com/photo-1503387762-592deb58ef4e?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
    {"slug": "supply-chain-finance", "key": "supply_chain_finance", "title": "Supply Chain Finance", "tagline": "Anchor-linked vendor & dealer funding",
     "rate_from": 10.25, "tenure_max": 12, "min_amount": 1000000, "max_amount": 500000000,
     "hero_image": "https://images.unsplash.com/photo-1553413077-190dd305871c?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
    {"slug": "invoice-discounting", "key": "invoice_discounting", "title": "Invoice Discounting", "tagline": "Advance against approved receivables",
     "rate_from": 10.5, "tenure_max": 6, "min_amount": 500000, "max_amount": 200000000,
     "hero_image": "https://images.unsplash.com/photo-1554224154-22dec7ec8818?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
    {"slug": "loan-against-securities", "key": "loan_against_securities", "title": "Loan Against Securities", "tagline": "Instant funding against MFs, shares, bonds",
     "rate_from": 9.0, "tenure_max": 36, "min_amount": 500000, "max_amount": 500000000,
     "hero_image": "https://images.unsplash.com/photo-1611974789855-9c2a0a7236a3?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
    {"slug": "structured-finance", "key": "structured_finance", "title": "Structured Finance", "tagline": "Custom solutions with warehousing lenders",
     "rate_from": 12.0, "tenure_max": 60, "min_amount": 50000000, "max_amount": 10000000000,
     "hero_image": "https://images.unsplash.com/photo-1454165804606-c3d57bc86b40?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
    {"slug": "private-credit", "key": "private_credit", "title": "Private Credit", "tagline": "Non-bank strategic capital",
     "rate_from": 14.0, "tenure_max": 60, "min_amount": 100000000, "max_amount": 20000000000,
     "hero_image": "https://images.unsplash.com/photo-1579532537598-459ecdaf39cc?crop=entropy&cs=srgb&fm=jpg&q=85&w=1200"},
]


@api.get("/public/products")
async def public_products():
    return PUBLIC_PRODUCTS


@api.get("/public/products/{slug}")
async def public_product(slug: str):
    p = next((x for x in PUBLIC_PRODUCTS if x["slug"] == slug), None)
    if not p: raise HTTPException(404, "Product not found")
    # Include an indicative lender panel
    lenders = await db.lenders.find({"products": p["key"], "active": True}, {"_id": 0, "name": 1, "roi_min": 1, "roi_max": 1, "lender_type": 1, "tat_days": 1}).limit(8).to_list(8)
    return {**p, "lenders": lenders}


@api.post("/public/apply")
async def public_apply(request: Request):
    """Public lead-capture endpoint used by the marketing site — no auth."""
    body = await request.json()
    if not body.get("name") or not body.get("mobile"):
        raise HTTPException(422, "name and mobile required")
    product_slug = body.get("product") or ""
    product_key = next((p["key"] for p in PUBLIC_PRODUCTS if p["slug"] == product_slug), "business_loan")
    lead_uid = await gen_uid(db, "lead")
    now_iso = datetime.now(timezone.utc).isoformat()
    # Default assign to a sales_agent (round-robin, simple)
    agents = await db.employees.find({"role": "sales_agent", "active": True}, {"_id": 0, "employee_uid": 1}).to_list(20)
    assigned = None
    if agents:
        n = await db.counters.find_one_and_update({"_id": "public_apply_rr"}, {"$inc": {"seq": 1}}, upsert=True, return_document=True)
        assigned = agents[(n["seq"] - 1) % len(agents)]["employee_uid"]
    doc = {
        "lead_uid": lead_uid, "source": "website", "source_detail": "public_apply",
        "campaign": body.get("campaign", ""), "referral": "",
        "channel_partner_uid": None,
        "assigned_to": assigned, "original_owner": assigned,
        "borrower_type": body.get("borrower_type", "business"),
        "name": body.get("name"), "company": body.get("company", ""),
        "mobile": body.get("mobile"), "email": body.get("email", ""),
        "city": body.get("city", ""), "state": body.get("state", ""),
        "product": product_key,
        "approx_requirement": float(body.get("amount") or 0),
        "notes": body.get("notes", "") + (f"\n\nTenure ask: {body.get('tenure_months')}m" if body.get('tenure_months') else ""),
        "stage": "new_lead", "priority": "hot", "probability": 40,
        "expected_closure": None, "rejection_reason": None,
        "client_uid": None, "converted": False, "duplicate_of": None,
        "created_at": now_iso,
    }
    await db.leads.insert_one(doc)
    doc.pop("_id", None)
    # If the user is logged in as a customer, link their user_id and echo lead_uid so their dashboard sees it
    try:
        u = await get_current_user(request, db)
        await db.leads.update_one({"lead_uid": lead_uid}, {"$set": {"customer_user_id": u["user_id"], "email": u["email"] or doc["email"]}})
    except Exception:
        pass
    await db.audit_logs.insert_one({
        "audit_id": f"aud_{uuid.uuid4().hex[:12]}", "actor_uid": "public",
        "actor_name": "Public site", "entity_type": "lead", "entity_id": lead_uid,
        "action": "public_apply", "before": None, "after": doc, "at": now_iso,
    })
    return {"ok": True, "lead_uid": lead_uid}


@api.get("/customer/me")
async def customer_me(request: Request):
    user = await require_user(request)
    if user.get("role") != "customer":
        raise HTTPException(403, "Customer portal only")
    leads = await db.leads.find({"$or": [{"customer_user_id": user["user_id"]}, {"email": user["email"]}]}, {"_id": 0}).to_list(200)
    client_uids = list({l.get("client_uid") for l in leads if l.get("client_uid")})
    emails = [user["email"]]
    clients = await db.clients.find({"$or": [{"client_uid": {"$in": client_uids}}, {"email": {"$in": emails}}]}, {"_id": 0}).to_list(50)
    case_client_uids = [c["client_uid"] for c in clients]
    cases = await db.cases.find({"client_uid": {"$in": case_client_uids}}, {"_id": 0}).to_list(100)
    # sanitize confidential fields
    for c in cases:
        for k in ["credit_owner", "expected_revenue", "actual_revenue"]:
            c.pop(k, None)
    return {"user": user, "leads": leads, "clients": clients, "cases": cases}



@api.post("/users/invite")
async def invite_user(request: Request):
    actor = await require_user(request)
    if actor["role"] != "super_admin":
        raise HTTPException(403, "Admin only")
    body = await request.json()
    email = (body.get("email") or "").strip().lower()
    if not email:
        raise HTTPException(422, "email required")
    role = body.get("role", "sales_agent")
    name = body.get("name") or email.split("@")[0].title()
    existing = await db.users.find_one({"email": email}, {"_id": 0})
    if existing:
        raise HTTPException(409, "User already exists")
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    emp_uid = await gen_uid(db, "employee") if role != "channel_partner" else None
    now_iso = datetime.now(timezone.utc).isoformat()
    doc = {
        "user_id": user_id, "email": email, "name": name, "picture": "",
        "role": role, "employee_uid": emp_uid,
        "partner_uid": body.get("partner_uid") if role == "channel_partner" else None,
        "active": True, "invited_by": actor["user_id"],
        "created_at": now_iso,
    }
    await db.users.insert_one(doc)
    doc.pop("_id", None)
    if emp_uid:
        await db.employees.insert_one({
            "employee_uid": emp_uid, "user_id": user_id, "email": email, "name": name,
            "role": role, "manager_uid": body.get("manager_uid"),
            "joining_date": now_iso, "ctc_monthly": float(body.get("ctc_monthly", 0)),
            "target_multiplier": 3.0, "revenue_target": 0, "disbursement_target": 0,
            "login_target": 0, "sanction_target": 0, "active": True, "created_at": now_iso,
        })
    await audit(actor, "user", user_id, "invited", None, doc)
    return doc


# ============== LEAD BULK IMPORT ==============
@api.post("/leads/import")
async def import_leads(request: Request):
    user = await require_user(request)
    body = await request.json()
    rows = body.get("rows") or []
    if not isinstance(rows, list) or not rows:
        raise HTTPException(422, "rows[] required")

    imported, duplicates, errors = [], [], []
    for i, r in enumerate(rows):
        try:
            mobile = str(r.get("mobile", "")).strip()
            email = str(r.get("email", "")).strip().lower()
            pan = str(r.get("pan", "")).strip().upper()
            gstin = str(r.get("gstin", "")).strip().upper()
            company = str(r.get("company", "")).strip()
            or_filters = []
            if mobile: or_filters.append({"mobile": mobile})
            if email: or_filters.append({"email": email})
            if pan: or_filters.append({"pan": pan})
            if gstin: or_filters.append({"gstin": gstin})
            if company: or_filters.append({"company": {"$regex": f"^{company}$", "$options": "i"}})
            dupe = None
            if or_filters:
                dupe_lead = await db.leads.find_one({"$or": or_filters}, {"_id": 0, "lead_uid": 1, "name": 1})
                dupe_client = await db.clients.find_one({"$or": or_filters}, {"_id": 0, "client_uid": 1, "name": 1})
                dupe = dupe_client or dupe_lead
            if dupe:
                duplicates.append({"row": i, "input": r, "match": dupe})
                continue
            if not r.get("name"):
                errors.append({"row": i, "error": "name required"})
                continue
            uid = await gen_uid(db, "lead")
            doc = {
                "lead_uid": uid, "source": r.get("source", "import"),
                "source_detail": r.get("source_detail", body.get("batch_id", "")),
                "campaign": r.get("campaign", ""), "referral": "",
                "channel_partner_uid": r.get("channel_partner_uid"),
                "assigned_to": r.get("assigned_to") or user.get("employee_uid"),
                "original_owner": r.get("assigned_to") or user.get("employee_uid"),
                "borrower_type": r.get("borrower_type", "business"),
                "name": r.get("name"), "company": company,
                "mobile": mobile, "email": email,
                "city": r.get("city", ""), "state": r.get("state", ""),
                "product": r.get("product", "business_loan"),
                "approx_requirement": float(r.get("approx_requirement") or 0),
                "notes": r.get("notes", ""),
                "stage": "new_lead", "priority": r.get("priority", "warm"),
                "probability": 30, "expected_closure": None,
                "rejection_reason": None, "client_uid": None,
                "converted": False, "duplicate_of": None,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "pan": pan, "gstin": gstin,
            }
            await db.leads.insert_one(doc)
            doc.pop("_id", None)
            imported.append(doc)
        except Exception as e:
            errors.append({"row": i, "error": str(e)})

    batch_id = f"IMP-{uuid.uuid4().hex[:8]}"
    await audit(user, "lead_import", batch_id, "batch",
                None, {"imported": len(imported), "duplicates": len(duplicates), "errors": len(errors)})
    return {"batch_id": batch_id, "imported": imported, "duplicates": duplicates, "errors": errors,
            "counts": {"imported": len(imported), "duplicates": len(duplicates), "errors": len(errors)}}


# ============== RENEWAL RADAR ==============
@api.get("/renewals")
async def renewals(request: Request):
    user = await require_user(request)
    scope = await scope_query(user, "cases")
    now = datetime.now(timezone.utc)
    query = {"stage": "fully_disbursed", **scope}
    cases = await db.cases.find(query, {"_id": 0}).to_list(500)
    rows = []
    for c in cases:
        # first disbursement date as loan_start
        first = await db.disbursements.find_one({"case_uid": c["case_uid"]}, {"_id": 0}, sort=[("disbursement_date", 1)])
        if not first: continue
        try:
            start = datetime.fromisoformat(first["disbursement_date"])
        except Exception:
            continue
        maturity = start + timedelta(days=int(c.get("tenure_months", 60)) * 30)
        days_to = (maturity - now).days
        if days_to > 180:  # only surface 6-month radar
            continue
        client = await db.clients.find_one({"client_uid": c["client_uid"]}, {"_id": 0, "name": 1, "company": 1, "mobile": 1, "relationship_manager": 1})
        bucket = "30" if days_to <= 30 else "60" if days_to <= 60 else "90" if days_to <= 90 else "180"
        rows.append({
            "case_uid": c["case_uid"], "client_uid": c["client_uid"],
            "client_name": (client or {}).get("name"), "company": (client or {}).get("company"),
            "mobile": (client or {}).get("mobile"), "rm": (client or {}).get("relationship_manager"),
            "product": c["product"], "disbursed_amount": c.get("disbursed_amount", 0),
            "roi": c.get("expected_roi", 0), "tenure_months": c.get("tenure_months", 0),
            "loan_start": first["disbursement_date"], "maturity_date": maturity.isoformat(),
            "days_to_maturity": days_to, "bucket": bucket,
        })
    rows.sort(key=lambda r: r["days_to_maturity"])
    return rows


# ============== SCHEDULED CRONS ==============
async def _process_hot_lead_followups():
    now = datetime.now(timezone.utc)
    cutoff_24h = (now - timedelta(hours=24)).isoformat()
    hot = await db.leads.find({
        "priority": "hot",
        "converted": {"$ne": True},
        "stage": {"$nin": ["closed", "lost", "rejected", "not_interested"]},
    }, {"_id": 0}).to_list(500)
    notified = 0
    app_url = os.environ.get("APP_PUBLIC_URL", "").rstrip("/")
    for lead in hot:
        latest = await db.activities.find_one(
            {"entity_type": "lead", "entity_id": lead["lead_uid"]},
            {"_id": 0, "created_at": 1}, sort=[("created_at", -1)]
        )
        last_touch = (latest or {}).get("created_at") or lead.get("created_at")
        if not last_touch or last_touch > cutoff_24h:
            continue
        owner_emp = await db.employees.find_one({"employee_uid": lead.get("assigned_to")}, {"_id": 0})
        if not owner_emp:
            continue
        manager_emp = None
        if owner_emp.get("manager_uid"):
            manager_emp = await db.employees.find_one({"employee_uid": owner_emp["manager_uid"]}, {"_id": 0})
        title = f"Hot lead going cold: {lead['name']}"
        body_msg = f"No activity in 24+ hours on hot lead {lead['lead_uid']}. Call or WhatsApp today."
        link = f"{app_url}/leads/{lead['lead_uid']}" if app_url else None
        for u in filter(None, [owner_emp, manager_emp]):
            await notify(db, u.get("user_id"), title, body_msg, link, "warning",
                         email=u.get("email"), mobile=None)
            notified += 1
        await db.tasks.insert_one({
            "task_id": await gen_uid(db, "task"),
            "title": f"[Auto] Re-engage hot lead {lead['lead_uid']}",
            "description": "Hot lead has been silent for 24+ hours. Call/WhatsApp today.",
            "owner_uid": lead.get("assigned_to"),
            "created_by": "system", "case_uid": None, "client_uid": None,
            "lead_uid": lead["lead_uid"], "application_uid": None,
            "priority": "urgent",
            "due_date": (now + timedelta(hours=6)).isoformat(),
            "status": "open", "origin": "auto_followup",
            "created_at": now.isoformat(),
        })
        await db.leads.update_one({"lead_uid": lead["lead_uid"]}, {"$set": {"stage": "escalated"}})
    return notified


async def _process_query_sla():
    """Escalate any lender query whose (created_at + lender.tat_days) < now."""
    now = datetime.now(timezone.utc)
    lender_tat = {l["lender_id"]: l.get("tat_days", 10) for l in await db.lenders.find({}, {"_id": 0, "lender_id": 1, "tat_days": 1}).to_list(200)}
    open_q = await db.lender_queries.find(
        {"status": {"$in": ["open", "awaiting_client", "awaiting_internal"]},
         "escalated": {"$ne": True}},
        {"_id": 0}
    ).to_list(500)
    escalated = 0
    app_url = os.environ.get("APP_PUBLIC_URL", "").rstrip("/")
    for q in open_q:
        app_doc = await db.applications.find_one({"application_uid": q.get("application_uid")}, {"_id": 0, "lender_id": 1})
        tat = lender_tat.get((app_doc or {}).get("lender_id"), 10)
        try:
            created = datetime.fromisoformat(q["created_at"])
        except Exception:
            continue
        due = created + timedelta(days=int(tat))
        if now < due:
            continue
        case = await db.cases.find_one({"case_uid": q["case_uid"]}, {"_id": 0})
        if not case: continue
        owner = await db.employees.find_one({"employee_uid": case.get("sales_owner")}, {"_id": 0})
        manager = None
        if owner and owner.get("manager_uid"):
            manager = await db.employees.find_one({"employee_uid": owner["manager_uid"]}, {"_id": 0})
        # Mark escalated + red flag on case
        await db.lender_queries.update_one({"query_id": q["query_id"]},
                                           {"$set": {"escalated": True, "escalated_at": now.isoformat()}})
        await db.cases.update_one({"case_uid": case["case_uid"]},
                                  {"$set": {"escalation_flag": "red"}})
        title = f"Query past TAT: {case['case_uid']}"
        body_msg = f"Lender query has been open beyond {tat}d TAT. Escalating for immediate action.\n\n\"{q['query_text']}\""
        link = f"{app_url}/cases/{case['case_uid']}" if app_url else None
        for u in filter(None, [owner, manager]):
            await notify(db, u.get("user_id"), title, body_msg, link, "warning",
                         email=u.get("email"), mobile=None)
            escalated += 1
    return escalated


async def _process_payout_batch():
    """Package all approved CP commissions + finance-approved incentives into one batch."""
    now = datetime.now(timezone.utc)
    period = now.strftime("%Y-%m")
    # eligible items
    cp = await db.cp_commissions.find(
        {"status": {"$in": ["approved", "accrued"]},
         "batch_id": {"$exists": False}}, {"_id": 0}
    ).to_list(1000)
    inc = await db.incentives.find(
        {"status": {"$in": ["manager_approved", "finance_approved", "payable"]},
         "batch_id": {"$exists": False}}, {"_id": 0}
    ).to_list(1000)
    if not cp and not inc:
        return {"batch_id": None, "period": period, "cp_count": 0, "incentive_count": 0, "total_amount": 0}
    batch_id = f"PO-{period}-{uuid.uuid4().hex[:6]}"

    cp_total = sum(x.get("payable_amount", 0) for x in cp)
    inc_total = sum((x.get("override_amount") or x.get("calculated_amount") or 0) for x in inc)

    # Build CSV
    lines = ["Beneficiary Type,ID,Case UID,Amount,TDS,Payable,Period,Reference"]
    for x in cp:
        lines.append(f"channel_partner,{x['partner_uid']},{x['case_uid']},{x['commission_amount']:.2f},{x.get('tds_amount', 0):.2f},{x.get('payable_amount', 0):.2f},{period},{x['commission_id']}")
    for x in inc:
        payable = (x.get("override_amount") or x.get("calculated_amount") or 0)
        lines.append(f"employee,{x['employee_uid']},,{payable:.2f},0,{payable:.2f},{x.get('period', period)},{x['incentive_id']}")
    csv_body = "\n".join(lines) + "\n"

    batch_doc = {
        "batch_id": batch_id, "period": period,
        "cp_count": len(cp), "incentive_count": len(inc),
        "cp_total": cp_total, "incentive_total": inc_total,
        "total_amount": cp_total + inc_total,
        "csv": csv_body, "status": "ready",
        "created_at": now.isoformat(),
    }
    await db.payout_batches.insert_one(batch_doc)
    batch_doc.pop("_id", None)

    # Mark items batched
    for x in cp:
        await db.cp_commissions.update_one({"commission_id": x["commission_id"]},
                                           {"$set": {"batch_id": batch_id}})
    for x in inc:
        await db.incentives.update_one({"incentive_id": x["incentive_id"]},
                                       {"$set": {"batch_id": batch_id}})

    # Notify finance team
    finance_users = await db.users.find({"role": "finance"}, {"_id": 0}).to_list(20)
    app_url = os.environ.get("APP_PUBLIC_URL", "").rstrip("/")
    for u in finance_users:
        await notify(db, u["user_id"],
                     f"Payout batch {batch_id} ready",
                     f"{len(cp)} channel partner commissions and {len(inc)} employee incentives — total {cp_total + inc_total:,.0f}. Download CSV in CorpZo → Payouts.",
                     f"{app_url}/payouts" if app_url else None, "info",
                     email=u.get("email"))
    return batch_doc


# ============== DOC DEFICIENCY → CORPZO ADVISORY OPPORTUNITIES ==============
DOC_CATEGORIES_REQUIRED = ["KYC", "Corporate", "Financial", "Banking", "GST/Tax", "Existing Loans", "Security/Collateral", "Legal"]

ADVISORY_SERVICES = {
    "KYC":                 {"name": "KYC & Compliance Advisory",          "fee": 5000,  "sla_days": 7},
    "Corporate":           {"name": "Company Secretarial & ROC Filings",  "fee": 15000, "sla_days": 14},
    "Financial":           {"name": "Financial Statement Prep & Audit",   "fee": 25000, "sla_days": 21},
    "Banking":             {"name": "Banking Advisory & Account Setup",   "fee": 10000, "sla_days": 7},
    "GST/Tax":             {"name": "GST Registration & Tax Advisory",    "fee": 15000, "sla_days": 14},
    "Existing Loans":      {"name": "Debt Consolidation & Refinance Advisory", "fee": 25000, "sla_days": 21},
    "Security/Collateral": {"name": "Property Valuation & Title Search",  "fee": 15000, "sla_days": 14},
    "Legal":               {"name": "Legal Documentation & Vetting",      "fee": 20000, "sla_days": 14},
}


@api.get("/cases/{case_uid}/doc-deficiency")
async def doc_deficiency(case_uid: str, request: Request):
    await require_user(request)
    case = await db.cases.find_one({"case_uid": case_uid}, {"_id": 0})
    if not case:
        raise HTTPException(404, "Case not found")
    present = set(await db.documents.distinct("category", {"case_uid": case_uid}))
    existing_opps = await db.opportunities.find(
        {"source_case_uid": case_uid}, {"deficient_doc_category": 1, "_id": 0}
    ).to_list(50)
    already = {o.get("deficient_doc_category") for o in existing_opps}
    missing = []
    for cat in DOC_CATEGORIES_REQUIRED:
        if cat in present:
            continue
        svc = ADVISORY_SERVICES.get(cat, {"name": f"{cat} Advisory", "fee": 5000, "sla_days": 7})
        missing.append({
            "category": cat, "service_name": svc["name"], "estimated_fee": svc["fee"],
            "sla_days": svc["sla_days"], "already_opportunity": cat in already,
        })
    present_req = [c for c in DOC_CATEGORIES_REQUIRED if c in present]
    return {
        "case_uid": case_uid,
        "present": sorted(list(present)),
        "missing": missing,
        "required_count": len(DOC_CATEGORIES_REQUIRED),
        "present_count": len(present_req),
    }


@api.post("/cases/{case_uid}/opportunities")
async def create_opportunity(case_uid: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    category = body.get("category")
    if not category:
        raise HTTPException(400, "category is required")
    case = await db.cases.find_one({"case_uid": case_uid}, {"_id": 0})
    if not case:
        raise HTTPException(404, "Case not found")
    dup = await db.opportunities.find_one({
        "source_case_uid": case_uid, "deficient_doc_category": category,
        "status": {"$ne": "dropped"}
    })
    if dup:
        raise HTTPException(400, f"Opportunity for {category} already exists on this case")
    svc = ADVISORY_SERVICES.get(category, {"name": f"{category} Advisory", "fee": 5000, "sla_days": 7})
    now = datetime.now(timezone.utc)
    opp_uid = await gen_uid(db, "opportunity")
    client = await db.clients.find_one({"client_uid": case.get("client_uid")}, {"_id": 0}) or {}
    doc = {
        "opportunity_uid": opp_uid,
        "source_case_uid": case_uid,
        "source_client_uid": case.get("client_uid"),
        "client_name": client.get("name") or client.get("company"),
        "client_mobile": client.get("mobile"),
        "deficient_doc_category": category,
        "service_name": svc["name"],
        "estimated_fee": body.get("estimated_fee", svc["fee"]),
        "sla_days": svc["sla_days"],
        "status": "open",
        # Determine advisory-desk owner: global setting overrides case sales owner
        "assigned_to": (await db.settings.find_one({"_id": "advisory_desk"}) or {}).get("owner_employee_uid") or case.get("sales_owner"),
        "notes": body.get("notes", ""),
        "created_by": user["user_id"],
        "created_at": now.isoformat(),
    }
    await db.opportunities.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "opportunity", opp_uid, "created", None, doc)
    await push_activity("case", case_uid, "opportunity", user,
                        f"Advisory opportunity created: {svc['name']} (est ₹{svc['fee']:,})")
    if case.get("sales_owner"):
        owner_user = await db.users.find_one({"employee_uid": case["sales_owner"]}, {"_id": 0})
        if owner_user:
            app_url = os.environ.get("APP_PUBLIC_URL", "").rstrip("/")
            await notify(db, owner_user["user_id"],
                         "New advisory opportunity",
                         f"{svc['name']} (est ₹{svc['fee']:,}) — from case {case_uid}",
                         f"{app_url}/opportunities" if app_url else None, "info",
                         email=owner_user.get("email"))
    return doc


@api.get("/settings/advisory-desk")
async def get_advisory_desk(request: Request):
    await require_user(request)
    setting = await db.settings.find_one({"_id": "advisory_desk"}) or {}
    owner_uid = setting.get("owner_employee_uid")
    owner = None
    if owner_uid:
        owner = await db.employees.find_one({"employee_uid": owner_uid}, {"_id": 0})
    return {"owner_employee_uid": owner_uid, "owner": owner}


@api.put("/settings/advisory-desk")
async def set_advisory_desk(request: Request):
    user = await require_user(request)
    if user.get("role") not in ("super_admin", "business_head"):
        raise HTTPException(403, "Only super_admin/business_head can change advisory desk owner")
    body = await request.json()
    owner_uid = body.get("owner_employee_uid") or None
    if owner_uid:
        exists = await db.employees.find_one({"employee_uid": owner_uid}, {"_id": 0})
        if not exists:
            raise HTTPException(400, "Employee not found")
    before = await db.settings.find_one({"_id": "advisory_desk"}, {"_id": 0})
    await db.settings.update_one({"_id": "advisory_desk"},
                                 {"$set": {"owner_employee_uid": owner_uid,
                                           "updated_at": datetime.now(timezone.utc).isoformat(),
                                           "updated_by": user["user_id"]}},
                                 upsert=True)
    after = await db.settings.find_one({"_id": "advisory_desk"}, {"_id": 0})
    await audit(user, "setting", "advisory_desk", "updated", before, after)
    return after


@api.get("/opportunities")
async def list_opportunities(request: Request, status: Optional[str] = None):
    await require_user(request)
    q = {}
    if status:
        q["status"] = status
    rows = await db.opportunities.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return rows


@api.patch("/opportunities/{opportunity_uid}")
async def update_opportunity(opportunity_uid: str, request: Request):
    user = await require_user(request)
    body = await request.json()
    before = await db.opportunities.find_one({"opportunity_uid": opportunity_uid}, {"_id": 0})
    if not before:
        raise HTTPException(404, "Not found")
    allowed = {"status", "assigned_to", "notes", "estimated_fee"}
    updates = {k: v for k, v in body.items() if k in allowed}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.opportunities.update_one({"opportunity_uid": opportunity_uid}, {"$set": updates})
    after = await db.opportunities.find_one({"opportunity_uid": opportunity_uid}, {"_id": 0})
    await audit(user, "opportunity", opportunity_uid, "updated", before, after)
    return after


# ============== WEEKLY REPORT EXPORTS ==============
def _week_range():
    now = datetime.now(timezone.utc)
    ws = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    return ws, now


async def _weekly_data():
    ws, we = _week_range()
    ws_iso = ws.isoformat()
    leads = await db.leads.find({"created_at": {"$gte": ws_iso}}, {"_id": 0}).to_list(1000)
    cases = await db.cases.find({"created_at": {"$gte": ws_iso}}, {"_id": 0}).to_list(1000)
    sanc = await db.sanctions.find({"created_at": {"$gte": ws_iso}}, {"_id": 0}).to_list(1000)
    disb = await db.disbursements.find({"created_at": {"$gte": ws_iso}}, {"_id": 0}).to_list(1000)
    invs = await db.invoices.find({"created_at": {"$gte": ws_iso}}, {"_id": 0}).to_list(1000)
    pays = await db.payments.find({"created_at": {"$gte": ws_iso}}, {"_id": 0}).to_list(1000)
    pipeline = await db.cases.aggregate([
        {"$group": {"_id": "$stage", "count": {"$sum": 1},
                    "requested": {"$sum": "$requirement"},
                    "sanctioned": {"$sum": "$sanctioned_amount"},
                    "disbursed": {"$sum": "$disbursed_amount"}}},
        {"$sort": {"count": -1}},
    ]).to_list(50)
    return {
        "week_start": ws.strftime("%d %b %Y"),
        "week_end":   we.strftime("%d %b %Y"),
        "leads":         {"total": len(leads), "rows": leads},
        "cases":         {"total": len(cases), "rows": cases},
        "sanctions":     {"count": len(sanc), "amount": sum(s.get("amount", 0) for s in sanc), "rows": sanc},
        "disbursements": {"count": len(disb), "amount": sum(x.get("amount", 0) for x in disb), "rows": disb},
        "invoices":      {"count": len(invs), "amount": sum(i.get("gross_amount", 0) for i in invs), "rows": invs},
        "payments":      {"count": len(pays), "amount": sum(p.get("amount", 0) for p in pays), "rows": pays},
        "pipeline": pipeline,
    }


def _inr(n):
    return f"Rs {(n or 0):,.0f}"


@api.get("/reports/weekly.xlsx")
async def report_weekly_xlsx(request: Request):
    await require_user(request)
    d = await _weekly_data()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF", size=11)
    head_fill = PatternFill("solid", fgColor="1F5B4A")
    title_font = Font(bold=True, size=14, color="0F3D2E")

    def style_head(ws, cols):
        for c in cols:
            ws[c].font = head_font
            ws[c].fill = head_fill

    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"CorpZo Weekly Report — {d['week_start']} to {d['week_end']}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")
    ws["A3"] = "Metric"; ws["B3"] = "Count"; ws["C3"] = "Amount"
    style_head(ws, ["A3", "B3", "C3"])
    summary = [
        ("New leads",       d["leads"]["total"],         "—"),
        ("New cases",       d["cases"]["total"],         "—"),
        ("Sanctions",       d["sanctions"]["count"],     _inr(d["sanctions"]["amount"])),
        ("Disbursements",   d["disbursements"]["count"], _inr(d["disbursements"]["amount"])),
        ("Invoices raised", d["invoices"]["count"],      _inr(d["invoices"]["amount"])),
        ("Revenue booked",  d["payments"]["count"],      _inr(d["payments"]["amount"])),
    ]
    for i, (label, count, amt) in enumerate(summary, start=4):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=count)
        ws.cell(row=i, column=3, value=amt)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20

    ws2 = wb.create_sheet("Pipeline")
    ws2.append(["Stage", "Count", "Requested", "Sanctioned", "Disbursed"])
    style_head(ws2, ["A1", "B1", "C1", "D1", "E1"])
    for row in d["pipeline"]:
        ws2.append([str(row.get("_id", "")).replace("_", " ").title(), row.get("count", 0),
                    row.get("requested", 0), row.get("sanctioned", 0), row.get("disbursed", 0)])
    for col, w in zip("ABCDE", [24, 10, 16, 16, 16]):
        ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet("Sanctions")
    ws3.append(["Sanction UID", "Case UID", "Lender", "Amount", "Status", "Sanctioned On"])
    style_head(ws3, ["A1", "B1", "C1", "D1", "E1", "F1"])
    for s in d["sanctions"]["rows"]:
        ws3.append([s.get("sanction_uid"), s.get("case_uid"), s.get("lender_name", ""),
                    s.get("amount", 0), s.get("status", ""), s.get("sanctioned_on", "")])
    for col, w in zip("ABCDEF", [20, 20, 24, 16, 14, 18]):
        ws3.column_dimensions[col].width = w

    ws4 = wb.create_sheet("Disbursements")
    ws4.append(["Disbursement UID", "Case UID", "Lender", "Amount", "Disbursed On"])
    style_head(ws4, ["A1", "B1", "C1", "D1", "E1"])
    for x in d["disbursements"]["rows"]:
        ws4.append([x.get("disbursement_uid"), x.get("case_uid"), x.get("lender_name", ""),
                    x.get("amount", 0), x.get("disbursed_on", "")])
    for col, w in zip("ABCDE", [22, 20, 24, 16, 18]):
        ws4.column_dimensions[col].width = w

    ws5 = wb.create_sheet("Revenue")
    ws5.append(["Payment UID", "Invoice UID", "Amount", "Received On", "Mode"])
    style_head(ws5, ["A1", "B1", "C1", "D1", "E1"])
    for p in d["payments"]["rows"]:
        ws5.append([p.get("payment_uid"), p.get("invoice_uid"), p.get("amount", 0),
                    p.get("received_on", ""), p.get("mode", "")])
    for col, w in zip("ABCDE", [20, 20, 16, 18, 14]):
        ws5.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"corpzo-weekly-{d['week_start'].replace(' ', '_')}.xlsx"
    return FastResponse(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api.get("/reports/weekly.pdf")
async def report_weekly_pdf(request: Request):
    await require_user(request)
    d = await _weekly_data()
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm

    buf = io.BytesIO()
    pdf = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=16*mm, rightMargin=16*mm,
                            topMargin=16*mm, bottomMargin=16*mm)
    styles = getSampleStyleSheet()
    brand_dark = colors.HexColor("#0F3D2E")
    brand      = colors.HexColor("#1F5B4A")

    title = ParagraphStyle("t", parent=styles["Heading1"], fontSize=20, textColor=brand_dark, spaceAfter=4)
    sub   = ParagraphStyle("s", parent=styles["Normal"], fontSize=10, textColor=colors.grey, spaceAfter=14)
    h2    = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=13, textColor=brand, spaceBefore=10, spaceAfter=6)

    story = [
        Paragraph("CorpZo — Weekly Management Report", title),
        Paragraph(f"{d['week_start']} to {d['week_end']}", sub),
        Paragraph("Executive summary", h2),
    ]

    kpi = [["Metric", "Count", "Amount"],
           ["New leads", d["leads"]["total"], "—"],
           ["New cases", d["cases"]["total"], "—"],
           ["Sanctions", d["sanctions"]["count"], _inr(d["sanctions"]["amount"])],
           ["Disbursements", d["disbursements"]["count"], _inr(d["disbursements"]["amount"])],
           ["Invoices raised", d["invoices"]["count"], _inr(d["invoices"]["amount"])],
           ["Revenue booked", d["payments"]["count"], _inr(d["payments"]["amount"])]]
    t = Table(kpi, colWidths=[70*mm, 30*mm, 45*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN",      (1, 0), (-1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(t)

    story.append(Paragraph("Pipeline by stage", h2))
    pipeline = [["Stage", "Count", "Requested", "Sanctioned", "Disbursed"]]
    for r in d["pipeline"]:
        pipeline.append([str(r.get("_id", "")).replace("_", " ").title(), r.get("count", 0),
                         _inr(r.get("requested", 0)), _inr(r.get("sanctioned", 0)), _inr(r.get("disbursed", 0))])
    tp = Table(pipeline, colWidths=[45*mm, 18*mm, 28*mm, 28*mm, 28*mm])
    tp.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
    ]))
    story.append(tp)

    if d["sanctions"]["rows"]:
        story.append(Paragraph("Sanctions this week", h2))
        rows = [["Sanction UID", "Case UID", "Lender", "Amount", "Status"]]
        for s in d["sanctions"]["rows"][:20]:
            rows.append([s.get("sanction_uid", ""), s.get("case_uid", ""),
                         (s.get("lender_name", "") or "")[:22],
                         _inr(s.get("amount", 0)), s.get("status", "")])
        ts = Table(rows, colWidths=[32*mm, 30*mm, 45*mm, 28*mm, 22*mm])
        ts.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), brand),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ]))
        story.append(ts)

    story.append(Spacer(1, 8*mm))
    story.append(Paragraph("<font size=8 color='#888'>Generated by CorpZo Debt CRM · Confidential — internal use only</font>", styles["Normal"]))
    pdf.build(story)
    buf.seek(0)
    fname = f"corpzo-weekly-{d['week_start'].replace(' ', '_')}.pdf"
    return FastResponse(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api.post("/cron/query-sla")
async def cron_query_sla(request: Request):
    # Cron endpoints must ack 2xx immediately; enqueue/background the actual work.
    import hmac
    auth = request.headers.get("Authorization", "")
    expected = os.environ.get("WEBHOOK_CRON_SECRET", "")
    if not auth.startswith("Bearer ") or not expected or not hmac.compare_digest(auth[7:], expected):
        raise HTTPException(401, "Unauthorized")
    asyncio.create_task(_process_query_sla())
    return {"accepted": True}


@api.post("/cron/payout-batch")
async def cron_payout_batch(request: Request):
    # Cron endpoints must ack 2xx immediately; enqueue/background the actual work.
    import hmac
    auth = request.headers.get("Authorization", "")
    expected = os.environ.get("WEBHOOK_CRON_SECRET", "")
    if not auth.startswith("Bearer ") or not expected or not hmac.compare_digest(auth[7:], expected):
        raise HTTPException(401, "Unauthorized")
    asyncio.create_task(_process_payout_batch())
    return {"accepted": True}


# ============== PAYOUTS (Finance) ==============
@api.get("/payouts")
async def list_payouts(request: Request):
    user = await require_user(request)
    payout_perms = {"release_commissions", "mark_payout_paid"}
    if user["role"] not in ("super_admin", "business_head", "finance") \
       and not (effective_permissions(user) & payout_perms):
        raise HTTPException(403, "Missing permission: release_commissions / mark_payout_paid")
    rows = await db.payout_batches.find({}, {"_id": 0, "csv": 0}).sort("created_at", -1).to_list(50)
    return rows


@api.post("/payouts/run-now")
async def run_payout_now(request: Request):
    user = await require_user(request)
    if not has_permission(user, "release_commissions"):
        raise HTTPException(403, "Missing permission: release_commissions")
    result = await _process_payout_batch()
    await audit(user, "payout", result.get("batch_id", "-"), "run_now", None, {"period": result.get("period"), "total": result.get("total_amount")})
    return result


@api.get("/payouts/{batch_id}/csv")
async def download_payout_csv(batch_id: str, request: Request):
    user = await require_user(request)
    payout_perms = {"release_commissions", "mark_payout_paid"}
    if user["role"] not in ("super_admin", "business_head", "finance") \
       and not (effective_permissions(user) & payout_perms):
        raise HTTPException(403, "Missing permission: release_commissions / mark_payout_paid")
    doc = await db.payout_batches.find_one({"batch_id": batch_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Batch not found")
    return FastResponse(
        content=doc["csv"], media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{batch_id}.csv"'},
    )


@api.post("/payouts/{batch_id}/mark-paid")
async def mark_batch_paid(batch_id: str, request: Request):
    user = await require_user(request)
    if not has_permission(user, "mark_payout_paid"):
        raise HTTPException(403, "Missing permission: mark_payout_paid")
    before = await db.payout_batches.find_one({"batch_id": batch_id}, {"_id": 0, "csv": 0})
    if not before:
        raise HTTPException(404, "Not found")
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.payout_batches.update_one({"batch_id": batch_id}, {"$set": {"status": "paid", "paid_at": now_iso}})
    await db.cp_commissions.update_many({"batch_id": batch_id}, {"$set": {"status": "paid", "paid_on": now_iso}})
    await db.incentives.update_many({"batch_id": batch_id}, {"$set": {"status": "paid", "paid_on": now_iso}})
    await audit(user, "payout", batch_id, "marked_paid", before, {"status": "paid"})
    return {"ok": True}





@api.post("/cron/hot-leads")
async def cron_hot_leads(request: Request):
    # Cron endpoints must ack 2xx immediately; enqueue/background the actual work.
    import hmac
    auth = request.headers.get("Authorization", "")
    expected = os.environ.get("WEBHOOK_CRON_SECRET", "")
    if not auth.startswith("Bearer ") or not expected or not hmac.compare_digest(auth[7:], expected):
        raise HTTPException(401, "Unauthorized")
    import asyncio as _a
    _a.create_task(_process_hot_lead_followups())
    return {"accepted": True}


@api.get("/notifications")
async def list_notifications(request: Request):
    user = await require_user(request)
    rows = await db.notifications.find({"user_id": user["user_id"]}, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)
    return rows


@api.post("/notifications/{notification_id}/read")
async def mark_read(notification_id: str, request: Request):
    user = await require_user(request)
    await db.notifications.update_one({"notification_id": notification_id, "user_id": user["user_id"]}, {"$set": {"read": True}})
    return {"ok": True}


# ============== PARTNER CRM v2 ==============
@api.patch("/channel-partners/{partner_uid}")
async def update_channel_partner(partner_uid: str, request: Request):
    user = await require_user(request)
    if not has_permission(user, "edit_partner_target"):
        raise HTTPException(403, "Missing permission: edit_partner_target")
    body = await request.json()
    before = await db.channel_partners.find_one({"partner_uid": partner_uid}, {"_id": 0})
    if not before:
        raise HTTPException(404, "Partner not found")
    allowed = {"name", "entity_name", "pan", "gst", "status", "kyc_status", "agreement_signed",
               "products", "geography", "channel_manager_uid", "commission_structure",
               "mobile", "email", "city", "state", "bank_account",
               "monthly_target", "notes"}
    updates = {k: v for k, v in body.items() if k in allowed}
    if updates:
        await db.channel_partners.update_one({"partner_uid": partner_uid}, {"$set": updates})
    after = await db.channel_partners.find_one({"partner_uid": partner_uid}, {"_id": 0})
    await audit(user, "channel_partner", partner_uid, "updated", before, after)
    return after


@api.get("/partners/performance")
async def partner_performance(request: Request):
    """Aggregate KPIs, MTD numbers, per-partner rollup and 6-month trend."""
    user = await require_user(request)
    partner_perms = {"release_commissions", "create_partner", "edit_partner_target"}
    if user.get("role") not in ("super_admin", "business_head", "channel_manager", "finance") \
       and not (effective_permissions(user) & partner_perms):
        raise HTTPException(403, "Not allowed")

    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_start_iso = month_start.isoformat()
    six_start = (month_start - timedelta(days=185)).replace(day=1).isoformat()

    partners = await db.channel_partners.find({}, {"_id": 0}).to_list(1000)
    partner_map = {p["partner_uid"]: p for p in partners}

    # Leads per partner
    leads = await db.leads.find({"channel_partner_uid": {"$ne": None}}, {"_id": 0}).to_list(5000)
    lead_count = {}
    for l in leads:
        pu = l.get("channel_partner_uid")
        if pu:
            lead_count[pu] = lead_count.get(pu, 0) + 1

    # Cases per partner
    cases = await db.cases.find({"channel_partner_uid": {"$ne": None}}, {"_id": 0}).to_list(5000)
    case_by_partner = {}
    for c in cases:
        pu = c.get("channel_partner_uid")
        if pu:
            case_by_partner.setdefault(pu, []).append(c)

    # Disbursements — need to join to case → partner
    case_partner = {c["case_uid"]: c.get("channel_partner_uid") for c in cases}
    disbursements = await db.disbursements.find(
        {"case_uid": {"$in": list(case_partner.keys())}}, {"_id": 0}
    ).to_list(5000)

    # Commissions
    commissions = await db.cp_commissions.find({}, {"_id": 0}).to_list(5000)

    # per-partner stats
    rows = []
    mtd_disb_total = 0
    mtd_comm_total = 0
    overdue_payout_total = 0
    active_partners = 0

    # Precompute monthly trend buckets per partner (last 6 months incl current)
    trend_labels = []
    _p = month_start
    for _ in range(6):
        trend_labels.append(_p.strftime("%Y-%m"))
        _prev = _p.replace(day=1) - timedelta(days=1)
        _p = _prev.replace(day=1)
    trend_labels = list(reversed(trend_labels))

    disb_trend_by_partner: Dict[str, Dict[str, float]] = {}
    for d in disbursements:
        pu = case_partner.get(d.get("case_uid"))
        if not pu:
            continue
        dt = (d.get("disbursement_date") or d.get("created_at") or "")[:7]
        if dt in trend_labels:
            disb_trend_by_partner.setdefault(pu, {}).setdefault(dt, 0.0)
            disb_trend_by_partner[pu][dt] += float(d.get("amount", 0) or 0)

    for p in partners:
        pu = p["partner_uid"]
        if p.get("status") == "active":
            active_partners += 1

        p_disbs = [d for d in disbursements if case_partner.get(d.get("case_uid")) == pu]
        p_disbursed_total = sum(float(x.get("amount", 0) or 0) for x in p_disbs)
        p_disbursed_mtd = sum(
            float(x.get("amount", 0) or 0) for x in p_disbs
            if (x.get("disbursement_date") or x.get("created_at") or "") >= month_start_iso
        )

        p_comms = [c for c in commissions if c.get("partner_uid") == pu]
        p_comm_total = sum(float(x.get("commission_amount", 0) or 0) for x in p_comms)
        p_comm_mtd = sum(
            float(x.get("commission_amount", 0) or 0) for x in p_comms
            if (x.get("created_at") or "") >= month_start_iso
        )
        p_comm_payable = sum(
            float(x.get("payable_amount", 0) or 0) for x in p_comms
            if x.get("status") in ("accrued", "approved") and not x.get("batch_id")
        )
        p_comm_overdue = sum(
            float(x.get("payable_amount", 0) or 0) for x in p_comms
            if x.get("status") == "accrued" and (x.get("created_at") or "") < (now - timedelta(days=45)).isoformat()
            and not x.get("batch_id")
        )

        target = float(p.get("monthly_target", 0) or 0)
        attainment_pct = round((p_disbursed_mtd / target) * 100, 1) if target > 0 else None

        # Trend series aligned to labels
        trend = disb_trend_by_partner.get(pu, {})
        series = [round(trend.get(lbl, 0.0), 2) for lbl in trend_labels]

        rows.append({
            "partner_uid": pu,
            "channel_code": p.get("channel_code", ""),
            "name": p.get("name", ""),
            "status": p.get("status", ""),
            "city": p.get("city", ""),
            "mobile": p.get("mobile", ""),
            "email": p.get("email", ""),
            "products": p.get("products", []),
            "monthly_target": target,
            "leads": lead_count.get(pu, 0),
            "cases": len(case_by_partner.get(pu, [])),
            "disbursed_total": round(p_disbursed_total, 2),
            "disbursed_mtd": round(p_disbursed_mtd, 2),
            "commission_total": round(p_comm_total, 2),
            "commission_mtd": round(p_comm_mtd, 2),
            "commission_payable": round(p_comm_payable, 2),
            "commission_overdue": round(p_comm_overdue, 2),
            "attainment_pct": attainment_pct,
            "trend_labels": trend_labels,
            "trend_disbursed": series,
        })

        mtd_disb_total += p_disbursed_mtd
        mtd_comm_total += p_comm_mtd
        overdue_payout_total += p_comm_overdue

    rows.sort(key=lambda r: r["disbursed_mtd"], reverse=True)

    return {
        "kpis": {
            "active_partners": active_partners,
            "total_partners": len(partners),
            "mtd_disbursement": round(mtd_disb_total, 2),
            "mtd_commission_accrued": round(mtd_comm_total, 2),
            "overdue_payouts": round(overdue_payout_total, 2),
        },
        "period": month_start.strftime("%b %Y"),
        "rows": rows,
    }


# ============== CAM PDF EXPORT ==============
@api.get("/cases/{case_uid}/cam.pdf")
async def cam_pdf(case_uid: str, request: Request):
    await require_user(request)
    c = await db.cases.find_one({"case_uid": case_uid}, {"_id": 0})
    if not c:
        raise HTTPException(404, "Case not found")
    client_doc = await db.clients.find_one({"client_uid": c["client_uid"]}, {"_id": 0}) or {}
    a = await db.assessments.find_one({"case_uid": case_uid}, {"_id": 0}) or {}
    bureau = await db.bureau_checks.find({"case_uid": case_uid}, {"_id": 0}).sort("checked_at", -1).to_list(1)
    latest_bureau = bureau[0] if bureau else {}

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm

    buf = io.BytesIO()
    pdf = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=16*mm, rightMargin=16*mm,
                            topMargin=16*mm, bottomMargin=16*mm)
    ss = getSampleStyleSheet()
    ink = colors.HexColor("#0F1B2D")
    brand = colors.HexColor("#0B1F3A")
    coral = colors.HexColor("#FF6B4E")
    gold = colors.HexColor("#D89B00")

    title = ParagraphStyle("t", parent=ss["Heading1"], fontSize=20, textColor=ink, spaceAfter=2)
    sub = ParagraphStyle("s", parent=ss["Normal"], fontSize=10, textColor=colors.grey, spaceAfter=10)
    h2 = ParagraphStyle("h2", parent=ss["Heading2"], fontSize=13, textColor=brand, spaceBefore=10, spaceAfter=6)
    p = ParagraphStyle("p", parent=ss["Normal"], fontSize=10, textColor=ink, spaceAfter=4)

    def _fmt_inr(x):
        try:
            n = float(x or 0)
        except (ValueError, TypeError):
            return "—"
        if n >= 1_00_00_000: return f"₹{n/1_00_00_000:.2f} Cr"
        if n >= 1_00_000: return f"₹{n/1_00_000:.2f} L"
        return f"₹{n:,.0f}"

    story = [
        Paragraph("Credit Assessment Memo (CAM)", title),
        Paragraph(f"Case {c['case_uid']} · Client {c['client_uid']} · Prepared {a.get('prepared_at', '')[:10] or datetime.now(timezone.utc).strftime('%Y-%m-%d')}", sub),
        Paragraph("Snapshot", h2),
    ]

    snap = [
        ["Client", (client_doc.get("company_name") or client_doc.get("name") or "—")[:40],
         "PAN", client_doc.get("pan", "—")],
        ["Product", str(c.get("product", "—")).replace("_", " ").title(),
         "Stage", str(c.get("stage", "—")).replace("_", " ").title()],
        ["Requested", _fmt_inr(c.get("requirement")),
         "Doc %", f"{c.get('documentation_pct', 0)}%"],
        ["Bureau Score", latest_bureau.get("score", "—"),
         "Bureau Bureau", latest_bureau.get("bureau", "—")],
    ]
    t = Table(snap, colWidths=[26*mm, 60*mm, 26*mm, 60*mm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (0, -1), brand),
        ("TEXTCOLOR", (2, 0), (2, -1), brand),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t)

    def _kv_table(label, section):
        if not section:
            return None
        story.append(Paragraph(label, h2))
        data = [["Field", "Value"]]
        for k, v in section.items():
            if v in (None, "", 0):
                continue
            data.append([str(k).replace("_", " ").title(), str(v)])
        if len(data) == 1:
            return None
        tab = Table(data, colWidths=[70*mm, 100*mm])
        tab.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), brand),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))
        story.append(tab)
        return tab

    _kv_table("Business overview", a.get("overview"))
    _kv_table("Financial performance", a.get("financials"))
    _kv_table("Banking analysis", a.get("banking"))
    _kv_table("Key ratios", a.get("ratios"))

    positives = a.get("positives") or []
    concerns = a.get("concerns") or []
    if positives or concerns:
        story.append(Paragraph("Positives & Concerns", h2))
        pc = [["Positives", "Concerns"], [
            "\n".join(f"• {x}" for x in positives) or "—",
            "\n".join(f"• {x}" for x in concerns) or "—",
        ]]
        tpc = Table(pc, colWidths=[85*mm, 85*mm])
        tpc.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#E6F5EE")),
            ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#FDEDEA")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(tpc)

    flags = a.get("flags") or []
    if flags:
        story.append(Paragraph("Credit flags", h2))
        f_rows = [["Level", "Title"]]
        for fl in flags:
            f_rows.append([str(fl.get("level", "")).upper(), str(fl.get("title", ""))])
        tf = Table(f_rows, colWidths=[25*mm, 145*mm])
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), brand),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]
        for i, fl in enumerate(flags, start=1):
            lvl = str(fl.get("level", ""))
            bg = {"green": colors.HexColor("#DCF5E6"), "amber": colors.HexColor("#FDF0D5"),
                  "red": colors.HexColor("#FDE1DE")}.get(lvl, colors.white)
            style.append(("BACKGROUND", (0, i), (0, i), bg))
        tf.setStyle(TableStyle(style))
        story.append(tf)

    story.append(Paragraph("Recommendation", h2))
    rec = a.get("recommendation") or "—"
    story.append(Paragraph(f"<b>Indicative eligibility:</b> {_fmt_inr(a.get('indicative_eligibility'))}", p))
    story.append(Paragraph(f"<b>Recommendation:</b> {rec.replace('_', ' ').title()}", p))
    if a.get("analyst_comments"):
        story.append(Paragraph("<b>Analyst comments</b>", p))
        story.append(Paragraph(a.get("analyst_comments").replace("\n", "<br/>"), p))

    story.append(Spacer(1, 10*mm))
    story.append(Paragraph(f"<font color='#94A3B8'>Prepared by {a.get('prepared_by', '—')} · Generated {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}</font>", sub))

    pdf.build(story)
    buf.seek(0)
    return FastResponse(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="CAM-{case_uid}.pdf"'},
    )


# ============== CAM TEMPLATES ==============
@api.get("/cam-templates")
async def list_cam_templates(request: Request, product: Optional[str] = None):
    await require_user(request)
    q = {}
    if product: q["product"] = product
    rows = await db.cam_templates.find(q, {"_id": 0}).sort("updated_at", -1).limit(200).to_list(200)
    return rows


@api.post("/cam-templates")
async def create_cam_template(request: Request):
    user = await require_user(request)
    require_permission(user, "publish_cam_template")
    body = await request.json()
    name = (body.get("name") or "").strip()
    product = (body.get("product") or "other").strip()
    if not name:
        raise HTTPException(422, "name required")
    template_id = f"tpl_{uuid.uuid4().hex[:10]}"
    # keep only non-PII structural fields
    snapshot = {
        "overview": {"industry": (body.get("overview") or {}).get("industry", "")},
        "financials": body.get("financials", {}),
        "banking": body.get("banking", {}),
        "ratios": body.get("ratios", {}),
        "positives": body.get("positives", []),
        "concerns": body.get("concerns", []),
        "flags": [
            {"level": (x.get("level") or "green"), "title": (x.get("title") or "")}
            for x in (body.get("flags") or []) if isinstance(x, dict) and x.get("title")
        ],
        "recommendation": body.get("recommendation", ""),
        "analyst_comments": body.get("analyst_comments", ""),
    }
    now_iso = datetime.now(timezone.utc).isoformat()
    doc = {
        "template_id": template_id, "name": name[:120], "product": product,
        "snapshot": snapshot, "author_uid": user["user_id"], "author_name": user.get("name", ""),
        "created_at": now_iso, "updated_at": now_iso,
    }
    await db.cam_templates.insert_one(doc)
    doc.pop("_id", None)
    await audit(user, "cam_template", template_id, "created", None, {"name": name, "product": product})
    return doc


@api.delete("/cam-templates/{template_id}")
async def delete_cam_template(template_id: str, request: Request):
    user = await require_user(request)
    tpl = await db.cam_templates.find_one({"template_id": template_id}, {"_id": 0})
    if not tpl:
        raise HTTPException(404, "Template not found")
    # Owner or publisher permission required
    if tpl.get("author_uid") != user["user_id"] and not has_permission(user, "publish_cam_template"):
        raise HTTPException(403, "Not allowed")
    await db.cam_templates.delete_one({"template_id": template_id})
    await audit(user, "cam_template", template_id, "deleted", tpl, None)
    return {"ok": True}


# ============== PUBLIC — LENDERS + PARTNER APPLICATIONS ==============
PUBLIC_LENDER_DIRECTORY = [
    # Scheduled commercial banks - Private
    {"name": "HDFC Bank",              "type": "Private Bank",     "tier": "top"},
    {"name": "ICICI Bank",             "type": "Private Bank",     "tier": "top"},
    {"name": "Axis Bank",              "type": "Private Bank",     "tier": "top"},
    {"name": "Kotak Mahindra Bank",    "type": "Private Bank",     "tier": "top"},
    {"name": "IndusInd Bank",          "type": "Private Bank",     "tier": "top"},
    {"name": "Yes Bank",               "type": "Private Bank"},
    {"name": "IDFC First Bank",        "type": "Private Bank"},
    {"name": "Federal Bank",           "type": "Private Bank"},
    {"name": "RBL Bank",               "type": "Private Bank"},
    {"name": "DCB Bank",               "type": "Private Bank"},
    {"name": "Karnataka Bank",         "type": "Private Bank"},
    {"name": "South Indian Bank",      "type": "Private Bank"},
    {"name": "City Union Bank",        "type": "Private Bank"},
    {"name": "Karur Vysya Bank",       "type": "Private Bank"},
    {"name": "Tamilnad Mercantile",    "type": "Private Bank"},
    {"name": "Bandhan Bank",           "type": "Private Bank"},
    {"name": "CSB Bank",               "type": "Private Bank"},
    {"name": "Jammu & Kashmir Bank",   "type": "Private Bank"},
    # PSU banks
    {"name": "State Bank of India",    "type": "PSU Bank",         "tier": "top"},
    {"name": "Punjab National Bank",   "type": "PSU Bank"},
    {"name": "Bank of Baroda",         "type": "PSU Bank"},
    {"name": "Union Bank of India",    "type": "PSU Bank"},
    {"name": "Bank of India",          "type": "PSU Bank"},
    {"name": "Canara Bank",            "type": "PSU Bank"},
    {"name": "Indian Bank",            "type": "PSU Bank"},
    {"name": "Indian Overseas Bank",   "type": "PSU Bank"},
    {"name": "Central Bank of India",  "type": "PSU Bank"},
    {"name": "UCO Bank",               "type": "PSU Bank"},
    {"name": "Bank of Maharashtra",    "type": "PSU Bank"},
    {"name": "Punjab & Sind Bank",     "type": "PSU Bank"},
    # Foreign banks
    {"name": "Standard Chartered",     "type": "Foreign Bank"},
    {"name": "HSBC India",             "type": "Foreign Bank"},
    {"name": "DBS Bank India",         "type": "Foreign Bank"},
    {"name": "Citi India",             "type": "Foreign Bank"},
    {"name": "Deutsche Bank India",    "type": "Foreign Bank"},
    {"name": "Barclays India",         "type": "Foreign Bank"},
    {"name": "MUFG Bank India",        "type": "Foreign Bank"},
    # Small finance banks
    {"name": "AU Small Finance Bank",  "type": "Small Finance Bank"},
    {"name": "Ujjivan SFB",            "type": "Small Finance Bank"},
    {"name": "Equitas SFB",            "type": "Small Finance Bank"},
    {"name": "ESAF SFB",               "type": "Small Finance Bank"},
    {"name": "Suryoday SFB",           "type": "Small Finance Bank"},
    {"name": "Utkarsh SFB",            "type": "Small Finance Bank"},
    {"name": "Fincare SFB",            "type": "Small Finance Bank"},
    {"name": "Jana SFB",               "type": "Small Finance Bank"},
    {"name": "Capital SFB",            "type": "Small Finance Bank"},
    {"name": "North East SFB",         "type": "Small Finance Bank"},
    # Retail NBFCs
    {"name": "Bajaj Finserv",          "type": "NBFC",             "tier": "top"},
    {"name": "Tata Capital",           "type": "NBFC",             "tier": "top"},
    {"name": "Aditya Birla Capital",   "type": "NBFC"},
    {"name": "Piramal Capital",        "type": "NBFC"},
    {"name": "L&T Finance",            "type": "NBFC"},
    {"name": "Poonawalla Fincorp",     "type": "NBFC"},
    {"name": "Hero FinCorp",           "type": "NBFC"},
    {"name": "Cholamandalam",          "type": "NBFC"},
    {"name": "Mahindra Finance",       "type": "NBFC"},
    {"name": "Shriram Finance",        "type": "NBFC"},
    {"name": "Muthoot Finance",        "type": "NBFC"},
    {"name": "Manappuram Finance",     "type": "NBFC"},
    {"name": "Fullerton India",        "type": "NBFC"},
    {"name": "IIFL Finance",           "type": "NBFC"},
    {"name": "Edelweiss",              "type": "NBFC"},
    {"name": "JM Financial",           "type": "NBFC"},
    {"name": "Northern Arc",           "type": "NBFC"},
    {"name": "Vivriti Capital",        "type": "NBFC"},
    {"name": "U GRO Capital",          "type": "NBFC"},
    {"name": "Lendingkart",            "type": "NBFC"},
    {"name": "FlexiLoans",             "type": "NBFC"},
    {"name": "Indifi",                 "type": "NBFC"},
    {"name": "NeoGrowth",              "type": "NBFC"},
    {"name": "InCred",                 "type": "NBFC"},
    {"name": "Clix Capital",           "type": "NBFC"},
    {"name": "DMI Finance",            "type": "NBFC"},
    {"name": "Godrej Capital",         "type": "NBFC"},
    {"name": "SMFG India Credit",      "type": "NBFC"},
    {"name": "Ambit Finvest",          "type": "NBFC"},
    {"name": "OxyzoFin",               "type": "NBFC"},
    {"name": "Kinara Capital",         "type": "NBFC"},
    {"name": "Ugro Business",          "type": "NBFC"},
    {"name": "Aye Finance",            "type": "NBFC"},
    {"name": "Arth Digital",           "type": "NBFC"},
    {"name": "Protium Finance",        "type": "NBFC"},
    {"name": "Growth Source",          "type": "NBFC"},
    {"name": "Krazybee",               "type": "NBFC"},
    {"name": "Whizdm (KreditBee)",     "type": "NBFC"},
    {"name": "MoneyView",              "type": "NBFC"},
    {"name": "Prefr (Bigul)",          "type": "NBFC"},
    # Housing finance
    {"name": "HDFC Ltd (Housing)",     "type": "Housing Finance"},
    {"name": "LIC Housing Finance",    "type": "Housing Finance"},
    {"name": "PNB Housing",            "type": "Housing Finance"},
    {"name": "Bajaj Housing",          "type": "Housing Finance"},
    {"name": "Aavas Financiers",       "type": "Housing Finance"},
    {"name": "Aptus Value Housing",    "type": "Housing Finance"},
    {"name": "Home First Finance",     "type": "Housing Finance"},
    {"name": "Repco Home Finance",     "type": "Housing Finance"},
    {"name": "IIFL Home Finance",      "type": "Housing Finance"},
    {"name": "India Shelter",          "type": "Housing Finance"},
    {"name": "Motilal Oswal HF",       "type": "Housing Finance"},
    # Education / specialty
    {"name": "Avanse",                 "type": "Specialty NBFC"},
    {"name": "Auxilo",                 "type": "Specialty NBFC"},
    {"name": "HDFC Credila",           "type": "Specialty NBFC"},
    {"name": "InCred Education",       "type": "Specialty NBFC"},
    {"name": "MPower Financing",       "type": "Specialty NBFC"},
    {"name": "Prodigy Finance",        "type": "Specialty NBFC"},
    # Private credit / AIFs (institutional)
    {"name": "Kotak Investment",       "type": "Private Credit"},
    {"name": "ICICI Prudential AIF",   "type": "Private Credit"},
    {"name": "Nomura India Credit",    "type": "Private Credit"},
    {"name": "Neo Wealth Credit",      "type": "Private Credit"},
    {"name": "Motilal Oswal Alt",      "type": "Private Credit"},
    {"name": "Edelweiss Alt Credit",   "type": "Private Credit"},
    {"name": "360 ONE Credit",         "type": "Private Credit"},
    {"name": "InCred Alternatives",    "type": "Private Credit"},
    {"name": "Nippon India AIF",       "type": "Private Credit"},
    {"name": "Multiples Alternate",    "type": "Private Credit"},
]


@api.get("/public/lenders")
async def public_lenders():
    """Full lender directory for the public /banks page."""
    by_type: Dict[str, list] = {}
    for l in PUBLIC_LENDER_DIRECTORY:
        by_type.setdefault(l["type"], []).append(l["name"])
    return {
        "total": len(PUBLIC_LENDER_DIRECTORY),
        "by_type": [{"type": k, "count": len(v), "names": v} for k, v in by_type.items()],
        "all": PUBLIC_LENDER_DIRECTORY,
    }


@api.post("/public/become-partner")
async def public_become_partner(request: Request):
    """Public form: aspiring channel partner submits interest. Stored as a lead + partner_application."""
    body = await request.json()
    name = (body.get("name") or "").strip()
    mobile = (body.get("mobile") or "").strip()
    email = (body.get("email") or "").strip()
    if not name or not mobile:
        raise HTTPException(422, "name and mobile required")

    now = datetime.now(timezone.utc)
    app_id = f"PA-{now.strftime('%Y%m')}-{uuid.uuid4().hex[:6].upper()}"
    doc = {
        "application_id": app_id,
        "name": name, "mobile": mobile, "email": email,
        "city": (body.get("city") or "").strip(),
        "state": (body.get("state") or "").strip(),
        "current_business": (body.get("current_business") or "").strip(),
        "expected_volume": body.get("expected_volume", ""),
        "products": body.get("products", []),
        "message": (body.get("message") or "").strip()[:1000],
        "source": "become_partner_page",
        "status": "new",
        "created_at": now.isoformat(),
    }
    await db.partner_applications.insert_one(doc)
    doc.pop("_id", None)
    # Notify channel_manager team (best-effort)
    try:
        cms = await db.users.find({"role": "channel_manager"}, {"_id": 0, "email": 1, "name": 1}).to_list(20)
        for cm in cms:
            if cm.get("email"):
                await send_email(cm["email"],
                                 f"New partner interest: {name}",
                                 f"<p><b>{name}</b> · {mobile} · {email or '—'}</p>"
                                 f"<p>City {doc['city']} · Products {', '.join(doc['products']) or '—'}</p>"
                                 f"<pre>{doc['message']}</pre>")
    except Exception:
        pass
    return doc


app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup():
    try:
        init_storage()
        log.info("Storage initialized")
    except Exception as e:
        log.error(f"Storage init failed: {e}")
    try:
        await seed_demo(db)
        await seed_supplemental(db)
    except Exception as e:
        log.error(f"Seed failed: {e}")


@app.on_event("shutdown")
async def shutdown():
    client.close()
